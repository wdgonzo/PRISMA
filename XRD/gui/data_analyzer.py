#!/usr/bin/env python3
"""
XRD Data Analyzer
=================
Clean, focused GUI for analyzing and visualizing processed XRD data.
Replaces the old complex visualization interfaces with a streamlined approach.

Features:
- Load and browse Zarr datasets
- Create heatmaps and plots
- Export CSV data
- Compare multiple datasets
- Simple, intuitive interface

Author(s): William Gonzalez, Luke Davenport
Date: October 2025
Version: Beta 0.1
"""

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QWidget, QFileDialog, QComboBox, QMessageBox, QGridLayout, QListWidget,
    QListWidgetItem, QSplitter, QTextEdit, QGroupBox, QCheckBox, QSpinBox,
    QDoubleSpinBox, QFormLayout, QProgressBar, QTabWidget, QScrollArea, QLineEdit,
    QRadioButton, QButtonGroup, QDialog
)
from PyQt5 import QtGui, QtCore
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
import json
import os
import sys
import glob
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import seaborn as sns

# Import data processing components
from XRD.core.gsas_processing import XRDDataset
from XRD.processing.recipes import load_recipe_from_file
from XRD.visualization import data_visualization


class SettingsManager:
    """Manages persistent application settings using JSON storage."""

    def __init__(self, app_name="XRDAnalyzer"):
        self.app_name = app_name
        self.settings_file = self.get_settings_path()
        self.default_settings = self.get_default_settings()

    def get_settings_path(self):
        """Get the path to the settings file."""
        # Use the same directory as the script for simplicity
        script_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(script_dir, f"{self.app_name}_settings.json")

    def get_default_settings(self):
        """Return the default settings structure."""
        return {
            "version": "3.0",
            "window": {
                "size": [1400, 900],
                "position": [100, 100]
            },
            "analysis_defaults": {
                "measurement": "strain",
                "delta_enabled": False,
                "abs_enabled": False,
                "pct_enabled": False,
                "graphing_mode": "heatmap",
                "colormap": "icefire",
                "vmin": -0.005,
                "vmax": 0.005,
                "auto_scale": True,
                "range_type": "full",
                "start_time": 0.0,
                "end_time": 1000.0,
                "start_frame": 0,
                "end_frame": 999999
            },
            "batch_analysis": {
                "frame_start": 150,
                "frame_end": 160,
                "azimuth_ranges": [
                    {"name": "Near Loading", "lower": -5, "upper": 5},
                    {"name": "Transverse 1", "lower": 85, "upper": 95},
                    {"name": "Opposite Loading", "lower": 175, "upper": 185},
                    {"name": "Transverse 2", "lower": -95, "upper": -85}
                ]
            },
            "paths": {
                "last_zarr_folder": ""
            },
            "ui_preferences": {
                "gallery_view_mode": "grid"
            }
        }

    def load_settings(self):
        """Load settings from file, return defaults if file doesn't exist or is corrupted."""
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r') as f:
                    settings = json.load(f)
                    # Merge with defaults to handle missing keys
                    return self.merge_with_defaults(settings)
            else:
                print(f"Settings file not found, using defaults: {self.settings_file}")
                return self.default_settings.copy()
        except (json.JSONDecodeError, IOError) as e:
            print(f"Error loading settings, using defaults: {e}")
            return self.default_settings.copy()

    def save_settings(self, settings):
        """Save settings to file."""
        try:
            # Create backup of existing settings
            if os.path.exists(self.settings_file):
                backup_file = self.settings_file + ".backup"
                import shutil
                shutil.copy2(self.settings_file, backup_file)

            with open(self.settings_file, 'w') as f:
                json.dump(settings, f, indent=2)
            return True
        except IOError as e:
            print(f"Error saving settings: {e}")
            return False

    def merge_with_defaults(self, settings):
        """Merge loaded settings with defaults to handle missing keys."""
        def deep_merge(default, loaded):
            result = default.copy()
            for key, value in loaded.items():
                if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                    result[key] = deep_merge(result[key], value)
                else:
                    result[key] = value
            return result

        return deep_merge(self.default_settings, settings)

    def reset_to_defaults(self):
        """Reset settings to factory defaults."""
        return self.default_settings.copy()


class BulkLoadWorker(QThread):
    """Worker thread for loading multiple Zarr datasets."""

    progress_updated = pyqtSignal(int, str)  # progress_value, current_dataset_name
    dataset_loaded = pyqtSignal(str, object)  # dataset_path, dataset_object
    load_error = pyqtSignal(str, str)  # dataset_path, error_message
    finished = pyqtSignal(int, int)  # loaded_count, total_count

    def __init__(self, zarr_dirs):
        super().__init__()
        self.zarr_dirs = zarr_dirs
        self.should_stop = False

    def run(self):
        """Load datasets in background thread."""
        loaded_count = 0
        total_count = len(self.zarr_dirs)

        for i, zarr_dir in enumerate(self.zarr_dirs):
            if self.should_stop:
                break

            try:
                dataset_name = os.path.basename(zarr_dir)
                self.progress_updated.emit(i, dataset_name)

                # Load the dataset (simplified version of load_dataset logic)
                metadata_path = os.path.join(zarr_dir, "metadata.json")
                if not os.path.exists(metadata_path):
                    raise ValueError("No metadata.json found")

                with open(metadata_path, 'r') as f:
                    metadata = json.load(f)

                from XRD.core.gsas_processing import GSASParams, Stages, PeakParams, XRDDataset

                # Load parameters from metadata
                params_data = metadata.get('params', {})

                # Reconstruct active_peaks from metadata
                active_peaks = []
                if 'active_peaks' in params_data and params_data['active_peaks']:
                    # Load peaks in original order from metadata
                    for peak_data in params_data['active_peaks']:
                        peak = PeakParams(
                            name=peak_data.get('name', 'Unknown Peak'),
                            miller_index=peak_data.get('miller_index', '211'),
                            position=peak_data.get('position', 8.46),
                            limits=tuple(peak_data.get('limits', [8.2, 8.8]))
                        )
                        active_peaks.append(peak)
                else:
                    # Fallback: detect peaks from dataset shape - use only generic names
                    n_peaks = metadata['n_peaks'] if 'n_peaks' in metadata else 3
                    active_peaks = []

                    # Just create generic peaks without assuming anything about positions or ranges
                    for i in range(n_peaks):
                        peak = PeakParams(f"Unknown {i+1}", f"hkl{i+1}", 8.0, (7.5, 8.5))
                        active_peaks.append(peak)

                # Get stage from metadata
                stage_name = params_data.get('stage', 'CONT')
                stage = getattr(Stages, stage_name, Stages.CONT)

                # Create GSASParams with all required fields
                params = GSASParams(
                    home_dir="",
                    images_path="",
                    refs_path=None,
                    control_file="",
                    mask_file="",
                    intplot_export=False,
                    sample=params_data.get('sample', 'Unknown'),
                    setting=params_data.get('setting', 'Unknown'),
                    stage=stage,
                    notes="",
                    exposure=params_data.get('exposure', '019'),
                    active_peaks=active_peaks,
                    azimuths=tuple(params_data.get('azimuths', (-110, 110))),
                    frames=tuple(params_data.get('frames', (0, 100))),
                    spacing=params_data.get('spacing', 5),
                    step=1,
                    pixel_size=(172.0, 172.0),
                    wavelength=0.1726,
                    detector_size=(1475, 1679)
                )

                # Load dataset
                dataset = XRDDataset.load(zarr_dir, params)

                # Basic validation
                if dataset.data.shape[0] == 0:
                    raise ValueError("Dataset has no peaks")

                self.dataset_loaded.emit(zarr_dir, dataset)
                loaded_count += 1

            except Exception as e:
                self.load_error.emit(zarr_dir, str(e))

        self.finished.emit(loaded_count, total_count)

    def stop(self):
        """Stop the loading process."""
        self.should_stop = True


class BulkLoadProgressDialog(QDialog):
    """Progress dialog for bulk loading Zarr datasets."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Loading Zarr Datasets")
        self.setModal(True)
        self.setFixedSize(400, 200)

        # Layout
        layout = QVBoxLayout(self)

        # Progress info
        self.status_label = QLabel("Preparing to load datasets...")
        layout.addWidget(self.status_label)

        # Progress bar
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # Current dataset label
        self.current_dataset_label = QLabel("")
        self.current_dataset_label.setStyleSheet("font-size: 10px; color: #666;")
        layout.addWidget(self.current_dataset_label)

        # Error summary
        self.error_summary = QTextEdit()
        self.error_summary.setMaximumHeight(80)
        self.error_summary.setVisible(False)
        layout.addWidget(self.error_summary)

        # Buttons
        button_layout = QHBoxLayout()
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        self.close_btn = QPushButton("Close")
        self.close_btn.clicked.connect(self.accept)
        self.close_btn.setVisible(False)

        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.close_btn)
        layout.addLayout(button_layout)

        # Track errors
        self.errors = []

    def update_progress(self, current, total, current_dataset):
        """Update the progress display."""
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current + 1)  # +1 to match the status text
        self.status_label.setText(f"Loading {current + 1} of {total} datasets...")
        self.current_dataset_label.setText(f"Current: {current_dataset}")

    def add_error(self, dataset_path, error_message):
        """Add an error to the summary."""
        dataset_name = os.path.basename(dataset_path)
        self.errors.append(f"{dataset_name}: {error_message}")

        if not self.error_summary.isVisible():
            self.error_summary.setVisible(True)
            self.setFixedSize(400, 300)

        self.error_summary.setText("\n".join(self.errors))

    def loading_finished(self, loaded_count, total_count):
        """Handle loading completion."""
        # Ensure progress bar shows 100%
        self.progress_bar.setValue(total_count)

        if loaded_count == total_count:
            self.status_label.setText(f"Successfully loaded all {total_count} datasets!")
            # Make close button more prominent and auto-close after 3 seconds
            self.close_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
            QTimer.singleShot(3000, self.accept)
        else:
            failed_count = total_count - loaded_count
            self.status_label.setText(f"Loaded {loaded_count} of {total_count} datasets ({failed_count} failed)")

        self.cancel_btn.setVisible(False)
        self.close_btn.setVisible(True)
        self.current_dataset_label.setText("Loading complete.")


class DatasetListWidget(QListWidget):
    """Widget for managing loaded datasets."""

    def __init__(self):
        super().__init__()
        self.datasets = {}  # Store dataset objects
        self.setSelectionMode(QListWidget.ExtendedSelection)

    def add_dataset(self, dataset_path: str, dataset: XRDDataset):
        """Add a dataset to the list."""
        # Create display name
        basename = os.path.basename(dataset_path)
        display_name = f"{dataset.params.sample} - {dataset.params.stage.name}"
        if hasattr(dataset.params, 'active_peaks'):
            peak_names = [peak.miller_index for peak in dataset.params.active_peaks[:2]]
            display_name += f" - {'/'.join(peak_names)}"

        # Add to list
        item = QListWidgetItem(display_name)
        item.setData(Qt.UserRole, dataset_path)
        self.addItem(item)

        # Store dataset
        self.datasets[dataset_path] = dataset

    def get_selected_datasets(self):
        """Get currently selected datasets."""
        selected = []
        for item in self.selectedItems():
            path = item.data(Qt.UserRole)
            if path in self.datasets:
                selected.append(self.datasets[path])
        return selected

    def clear_datasets(self):
        """Clear all datasets."""
        self.clear()
        self.datasets.clear()


class PlotCanvas(FigureCanvas):
    """Matplotlib canvas for displaying plots."""

    def __init__(self, parent=None, width=10, height=8, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        super().__init__(self.fig)
        self.setParent(parent)
        self.selection_box = None  # Store selection box rectangle

    def clear_plot(self):
        """Clear the current plot."""
        self.fig.clear()
        self.selection_box = None
        self.draw()

    def plot_heatmap(self, data, xlabel="Azimuth", ylabel="Time/Depth",
                     colormap="viridis", vmin=None, vmax=None, cbar_label="", measurement_type="", azimuth_range=(-110, 110), spacing=5, title="", is_time_based=True):
        """Plot a heatmap matching the old visualization implementation."""
        self.fig.clear()

        # Set font size for this figure only (matching old visualization)
        import matplotlib as mpl
        original_fontsize = mpl.rcParams['font.size']
        mpl.rcParams.update({'font.size': 12})  # Reduced to match old visualization better

        ax = self.fig.add_subplot(111)

        # Use seaborn heatmap for proper data handling and orientation
        # data should be a pandas DataFrame or 2D array where:
        # - rows = time/depth (y-axis)
        # - columns = azimuth (x-axis)
        # Colorbar settings to match old visualization - shrink=1.0 for full height
        cbar_kwargs = {"shrink": 1.0, "label": cbar_label}

        if vmin is not None and vmax is not None:
            im = sns.heatmap(data, cmap=colormap, vmin=vmin, vmax=vmax, ax=ax, cbar_kws=cbar_kwargs)
        else:
            im = sns.heatmap(data, cmap=colormap, ax=ax, cbar_kws=cbar_kwargs)

        # Add title to the plot (matching old visualization positioning)
        if title:
            ax.set_title(title, y=1.05)

        # Configure colorbar appearance to match old visualization
        if hasattr(im, 'collections') and len(im.collections) > 0:
            cbar = im.collections[0].colorbar
            if cbar:
                cbar.outline.set_edgecolor('black')
                cbar.outline.set_linewidth(1.5)

        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)

        # Configure axis tick formatting to match old implementation
        self.configure_axis_formatting(ax, data, azimuth_range, is_time_based)

        # Set aspect ratio for square pixels (matching old implementation)
        if hasattr(data, 'index') and len(data.index) > 0:
            az_start, az_end = azimuth_range
            ratio = ((az_end - az_start) / spacing) / len(data.index)
            ax.set_aspect(ratio)

        # Use subplot adjust instead of tight_layout for precise control (matching old visualization)
        if measurement_type.startswith("strain"):
            self.fig.subplots_adjust(left=0.14, right=0.88, top=0.93, bottom=0.08)
        else:
            self.fig.subplots_adjust(left=0.14, right=0.93, top=0.93, bottom=0.08)

        self.draw()

        # Restore original font size
        mpl.rcParams.update({'font.size': original_fontsize})

    def configure_axis_formatting(self, ax, data, azimuth_range=(-110, 110), is_time_based=True):
        """Configure axis tick formatting to match old visualization implementation."""

        # X-axis formatting (azimuth angles)
        if hasattr(data, 'columns'):
            try:
                azimuths = np.array([float(col) for col in data.columns])
                az_min, az_max = azimuth_range

                # Calculate appropriate tick spacing to prevent overlap
                # Use 30° for standard ranges, 45° for wider ranges, always include 0
                angle_range = az_max - az_min
                if angle_range > 270:
                    tick_spacing = 60  # Full circle or near-full: 60° spacing
                elif angle_range > 180:
                    tick_spacing = 45  # Wide range: 45° spacing
                else:
                    tick_spacing = 30  # Standard range: 30° spacing

                # Build tick list, always including 0 if it's in range
                start = ((az_min + tick_spacing - 1) // tick_spacing) * tick_spacing
                end = (az_max // tick_spacing) * tick_spacing
                x_ticks = list(np.arange(start, end + 1, tick_spacing))

                # Ensure 0 is included if it's within the azimuth range
                if az_min <= 0 <= az_max and 0 not in x_ticks:
                    x_ticks.append(0)
                    x_ticks.sort()

                # Add final boundary tick at az_max if not already included
                if az_max not in x_ticks and abs(x_ticks[-1] - az_max) > tick_spacing / 2:
                    x_ticks.append(az_max)

                # Find closest data points to desired ticks and build labels
                x_tick_indices = []
                x_tick_labels = []
                for tick in x_ticks:
                    if tick >= azimuths.min() and tick <= azimuths.max():
                        idx = np.abs(azimuths - tick).argmin()
                        x_tick_indices.append(idx)
                        # Use the ideal tick value for the label, not the actual data value
                        x_tick_labels.append(f"{int(tick)}")

                ax.set_xticks(x_tick_indices)
                ax.set_xticklabels(x_tick_labels, rotation=0, fontsize=10)

                # Set x-axis limits with small margin for blank padding
                ax.set_xlim(-1.5, data.shape[1] + 1.5)

            except (ValueError, IndexError):
                pass

        # Y-axis formatting based on data type
        if hasattr(data, 'index') and len(data.index) > 0:
            try:
                if is_time_based:
                    # Time-based: 15 evenly spaced ticks (matching old visualizer)
                    y_ticks = np.linspace(0, len(data.index) - 1, 15, dtype=int)
                    ax.set_yticks(y_ticks)
                    ax.set_yticklabels([f"{data.index[i]:.2f}" for i in y_ticks])
                else:
                    # Depth-based: ticks every 20 μm (matching old visualizer)
                    y_ticks = np.arange(0, len(data.index), 20)
                    if len(y_ticks) > 0:
                        offset = len(data.index) - y_ticks[-1]
                        ax.set_yticks(np.flip(y_ticks + offset))
                        ax.set_yticklabels([f"{i:.0f}" for i in y_ticks])
            except (ValueError, IndexError):
                pass

    def draw_selection_box(self, frame_start, frame_end, azimuth_start, azimuth_end, data):
        """Draw a selection box overlay on the current plot."""
        if not self.fig.axes:
            return

        ax = self.fig.axes[0]

        # Clear existing selection box
        if self.selection_box:
            self.selection_box.remove()

        try:
            # Convert frame and azimuth ranges to plot coordinates
            azimuths = np.array([float(col) for col in data.columns])

            # Find closest azimuth indices
            azimuth_start_idx = np.abs(azimuths - azimuth_start).argmin()
            azimuth_end_idx = np.abs(azimuths - azimuth_end).argmin()

            # Ensure correct order
            if azimuth_start_idx > azimuth_end_idx:
                azimuth_start_idx, azimuth_end_idx = azimuth_end_idx, azimuth_start_idx

            # Frame indices (they match data index directly for the current display)
            frame_start_idx = None
            frame_end_idx = None

            # Find frame indices by matching against data index
            for i, idx_value in enumerate(data.index):
                if frame_start_idx is None and idx_value >= frame_start:
                    frame_start_idx = i
                if idx_value <= frame_end:
                    frame_end_idx = i

            if frame_start_idx is None:
                frame_start_idx = 0
            if frame_end_idx is None:
                frame_end_idx = len(data.index) - 1

            # Ensure correct order
            if frame_start_idx > frame_end_idx:
                frame_start_idx, frame_end_idx = frame_end_idx, frame_start_idx

            # Create rectangle (x, y, width, height)
            x = azimuth_start_idx - 0.5
            y = frame_start_idx - 0.5
            width = azimuth_end_idx - azimuth_start_idx + 1
            height = frame_end_idx - frame_start_idx + 1

            # Draw selection box
            from matplotlib.patches import Rectangle
            self.selection_box = Rectangle(
                (x, y), width, height,
                linewidth=2, edgecolor='yellow', facecolor='yellow', alpha=0.3
            )
            ax.add_patch(self.selection_box)

            # Redraw
            self.draw()

        except Exception as e:
            print(f"Error drawing selection box: {e}")

    def clear_selection_box(self):
        """Clear the selection box if it exists."""
        if self.selection_box:
            self.selection_box.remove()
            self.selection_box = None
            self.draw()


class AnalysisTab(QWidget):
    """Individual analysis tab with parameters and controls."""

    plotRequested = pyqtSignal(dict, str)  # Signal to generate plot

    def __init__(self, tab_name="Analysis", dataset=None):
        super().__init__()
        self.tab_name = tab_name
        self.dataset = dataset
        self.current_measurement = "strain"
        self.current_peak_idx = 0
        self.setup_ui()

    def setup_ui(self):
        """Setup the analysis tab UI."""
        # Create main layout for the tab
        main_layout = QVBoxLayout(self)

        # Create scroll area for all content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # Create content widget that will go inside the scroll area
        content_widget = QWidget()
        layout = QVBoxLayout(content_widget)

        # Tab name entry
        header_layout = QHBoxLayout()
        name_label = QLabel("Analysis Name:")
        self.name_entry = QLineEdit(self.tab_name)
        header_layout.addWidget(name_label)
        header_layout.addWidget(self.name_entry)
        header_layout.addStretch()
        layout.addLayout(header_layout)

        # Analysis parameters group
        params_group = QGroupBox("Analysis Parameters")
        params_layout = QFormLayout()

        # Base measurement selection (matching actual stored data)
        self.measurement_combo = QComboBox()
        base_measurements = ["strain", "d", "area", "sigma", "gamma", "pos"]
        self.measurement_combo.addItems(base_measurements)
        self.measurement_combo.currentTextChanged.connect(self.on_measurement_changed)

        # Delta, Abs, and PCT checkboxes (mutually exclusive)
        self.delta_check = QCheckBox("Apply delta (time-based change)")
        self.delta_check.toggled.connect(self.on_delta_toggled)

        self.abs_check = QCheckBox("Apply absolute value")
        self.abs_check.toggled.connect(self.on_abs_toggled)

        self.pct_check = QCheckBox("Apply percentage (measurement/ref)")
        self.pct_check.toggled.connect(self.on_pct_toggled)

        # Graphing mode
        self.graphing_mode_combo = QComboBox()
        self.graphing_mode_combo.addItems(["Standard", "Robust", "Robust L.L.", "Standard L.L."])
        self.graphing_mode_combo.currentTextChanged.connect(self.on_parameter_changed)

        self.peak_combo = QComboBox()
        self.peak_combo.currentIndexChanged.connect(self.on_peak_changed)

        self.colormap_combo = QComboBox()
        self.colormap_combo.addItems(["icefire", "viridis", "plasma", "inferno", "coolwarm", "RdBu_r"])

        self.vmin_spin = QDoubleSpinBox()
        self.vmin_spin.setRange(-10.0, 10.0)
        self.vmin_spin.setSingleStep(0.001)
        self.vmin_spin.setDecimals(4)
        self.vmin_spin.setValue(-0.005)

        self.vmax_spin = QDoubleSpinBox()
        self.vmax_spin.setRange(-10.0, 10.0)
        self.vmax_spin.setSingleStep(0.001)
        self.vmax_spin.setDecimals(4)
        self.vmax_spin.setValue(0.005)

        self.auto_scale_check = QCheckBox()
        self.auto_scale_check.setChecked(True)
        self.auto_scale_check.toggled.connect(self.on_auto_scale_toggled)

        params_layout.addRow("Measurement:", self.measurement_combo)
        params_layout.addRow("", self.delta_check)
        params_layout.addRow("", self.abs_check)
        params_layout.addRow("", self.pct_check)
        params_layout.addRow("Graphing Mode:", self.graphing_mode_combo)
        params_layout.addRow("Peak:", self.peak_combo)
        params_layout.addRow("Colormap:", self.colormap_combo)
        params_layout.addRow("Min Value:", self.vmin_spin)
        params_layout.addRow("Max Value:", self.vmax_spin)
        params_layout.addRow("Auto Scale:", self.auto_scale_check)

        params_group.setLayout(params_layout)
        layout.addWidget(params_group)

        # Frame/Time Range selection group
        range_group = QGroupBox("Frame/Time Range")
        range_layout = QFormLayout()

        # Radio buttons for range type selection
        self.range_button_group = QButtonGroup()

        self.full_range_radio = QRadioButton("Full Range")
        self.full_range_radio.setChecked(True)  # Default
        self.time_range_radio = QRadioButton("Time Range (minutes)")
        self.frame_range_radio = QRadioButton("Frame Range")

        self.range_button_group.addButton(self.full_range_radio, 0)
        self.range_button_group.addButton(self.time_range_radio, 1)
        self.range_button_group.addButton(self.frame_range_radio, 2)

        # Connect radio button changes
        self.range_button_group.buttonClicked.connect(self.on_range_type_changed)

        # Time range controls
        self.start_time_spin = QDoubleSpinBox()
        self.start_time_spin.setRange(0.0, 999.0)
        self.start_time_spin.setSingleStep(0.1)
        self.start_time_spin.setDecimals(2)
        self.start_time_spin.setValue(0.0)
        self.start_time_spin.setEnabled(False)

        self.end_time_spin = QDoubleSpinBox()
        self.end_time_spin.setRange(0.0, 999.0)
        self.end_time_spin.setSingleStep(0.1)
        self.end_time_spin.setDecimals(2)
        self.end_time_spin.setValue(10.0)
        self.end_time_spin.setEnabled(False)

        # Frame range controls
        self.start_frame_spin = QSpinBox()
        self.start_frame_spin.setRange(0, 999999)
        self.start_frame_spin.setValue(0)
        self.start_frame_spin.setEnabled(False)

        self.end_frame_spin = QSpinBox()
        self.end_frame_spin.setRange(0, 999999)
        self.end_frame_spin.setValue(1000)
        self.end_frame_spin.setEnabled(False)

        # Auto-detect button
        self.auto_detect_btn = QPushButton("Auto-detect Range")
        self.auto_detect_btn.clicked.connect(self.auto_detect_range)
        self.auto_detect_btn.setEnabled(False)

        # Add range controls to layout
        range_layout.addRow("", self.full_range_radio)
        range_layout.addRow("", self.time_range_radio)
        range_layout.addRow("Start Time:", self.start_time_spin)
        range_layout.addRow("End Time:", self.end_time_spin)
        range_layout.addRow("", self.frame_range_radio)
        range_layout.addRow("Start Frame:", self.start_frame_spin)
        range_layout.addRow("End Frame:", self.end_frame_spin)
        range_layout.addRow("", self.auto_detect_btn)

        range_group.setLayout(range_layout)
        layout.addWidget(range_group)

        # Selection Analysis group
        selection_group = QGroupBox("Selection Analysis")
        selection_layout = QFormLayout()

        # Frame range for selection
        self.selection_start_frame_spin = QSpinBox()
        self.selection_start_frame_spin.setRange(0, 999999)
        self.selection_start_frame_spin.setValue(150)

        self.selection_end_frame_spin = QSpinBox()
        self.selection_end_frame_spin.setRange(0, 999999)
        self.selection_end_frame_spin.setValue(160)

        # Azimuth range for selection
        self.selection_start_azimuth_spin = QDoubleSpinBox()
        self.selection_start_azimuth_spin.setRange(-360.0, 360.0)
        self.selection_start_azimuth_spin.setSingleStep(1.0)
        self.selection_start_azimuth_spin.setValue(-5.0)

        self.selection_end_azimuth_spin = QDoubleSpinBox()
        self.selection_end_azimuth_spin.setRange(-360.0, 360.0)
        self.selection_end_azimuth_spin.setSingleStep(1.0)
        self.selection_end_azimuth_spin.setValue(5.0)

        # Calculate button
        self.calculate_average_btn = QPushButton("Calculate Average")
        self.calculate_average_btn.clicked.connect(self.calculate_selection_average)

        # Connect range inputs to real-time updates
        self.selection_start_frame_spin.valueChanged.connect(self.update_selection_preview)
        self.selection_end_frame_spin.valueChanged.connect(self.update_selection_preview)
        self.selection_start_azimuth_spin.valueChanged.connect(self.update_selection_preview)
        self.selection_end_azimuth_spin.valueChanged.connect(self.update_selection_preview)

        # Average display
        self.selection_average_label = QLabel("Average of Selection: --")
        self.selection_average_label.setStyleSheet("font-weight: bold; color: #0066cc; background-color: #f0f8ff; padding: 5px; border: 1px solid #cce7ff;")

        # Add to layout
        selection_layout.addRow("Frame Start:", self.selection_start_frame_spin)
        selection_layout.addRow("Frame End:", self.selection_end_frame_spin)
        selection_layout.addRow("Azimuth Start (°):", self.selection_start_azimuth_spin)
        selection_layout.addRow("Azimuth End (°):", self.selection_end_azimuth_spin)
        selection_layout.addRow("", self.calculate_average_btn)
        selection_layout.addRow("", self.selection_average_label)

        selection_group.setLayout(selection_layout)
        layout.addWidget(selection_group)

        # Action buttons
        buttons_layout = QHBoxLayout()

        preview_btn = QPushButton("Preview")
        preview_btn.clicked.connect(self.request_preview)

        buttons_layout.addWidget(preview_btn)
        buttons_layout.addStretch()

        layout.addLayout(buttons_layout)
        layout.addStretch()

        # Set the content widget in the scroll area
        scroll_area.setWidget(content_widget)

        # Add the scroll area to the main layout
        main_layout.addWidget(scroll_area)

    def update_for_dataset(self, dataset):
        """Update tab for new dataset."""
        self.dataset = dataset
        if dataset:
            # Update peak combo with available peaks
            self.peak_combo.clear()

            if hasattr(dataset.params, 'active_peaks') and dataset.params.active_peaks:
                # Use defined peaks
                print(f"Populating peak combo with metadata peaks:")
                for i, peak in enumerate(dataset.params.active_peaks):
                    print(f"  Combo Index {i}: {peak.name} ({peak.miller_index}) at {peak.position}")
                    self.peak_combo.addItem(f"{peak.name} ({peak.miller_index})", i)
            else:
                # Fallback: detect peaks from dataset shape - use only generic names
                n_peaks = dataset.data.shape[0] if dataset.data is not None else 3

                # Just create generic peak names without assumptions
                for i in range(n_peaks):
                    self.peak_combo.addItem(f"Unknown {i+1} (hkl{i+1})", i)

                print(f"Populated peak combo with {n_peaks} detected peaks")

    def on_measurement_changed(self):
        self.update_current_measurement()

    def on_peak_changed(self):
        self.current_peak_idx = self.peak_combo.currentData() or 0

    def on_parameter_changed(self):
        pass  # For future parameter validation

    def on_auto_scale_toggled(self, checked):
        self.vmin_spin.setEnabled(not checked)
        self.vmax_spin.setEnabled(not checked)

    def on_delta_toggled(self, checked):
        """Handle delta checkbox toggle - make mutually exclusive with abs and pct."""
        if checked:
            if self.abs_check.isChecked():
                self.abs_check.setChecked(False)
            if self.pct_check.isChecked():
                self.pct_check.setChecked(False)
        self.update_current_measurement()

    def on_abs_toggled(self, checked):
        """Handle abs checkbox toggle - make mutually exclusive with delta and pct."""
        if checked:
            if self.delta_check.isChecked():
                self.delta_check.setChecked(False)
            if self.pct_check.isChecked():
                self.pct_check.setChecked(False)
        self.update_current_measurement()

    def on_pct_toggled(self, checked):
        """Handle pct checkbox toggle - make mutually exclusive with delta and abs."""
        if checked:
            if self.delta_check.isChecked():
                self.delta_check.setChecked(False)
            if self.abs_check.isChecked():
                self.abs_check.setChecked(False)
        self.update_current_measurement()

    def update_current_measurement(self):
        """Update current_measurement based on base measurement and checkboxes."""
        base = self.measurement_combo.currentText()
        if self.delta_check.isChecked():
            self.current_measurement = f"delta {base}"
        elif self.abs_check.isChecked():
            self.current_measurement = f"abs {base}"
        elif self.pct_check.isChecked():
            self.current_measurement = f"pct {base}"
        else:
            self.current_measurement = base

    def request_preview(self):
        if self.dataset:
            params = self.get_analysis_parameters()
            self.plotRequested.emit(params, "preview")

    def request_generation(self):
        if self.dataset:
            params = self.get_analysis_parameters()
            self.plotRequested.emit(params, "generate")

    def get_analysis_parameters(self):
        """Get current analysis parameters including range selection."""
        params = {
            'measurement': self.current_measurement,
            'mode': self.graphing_mode_combo.currentText(),
            'peak_idx': self.current_peak_idx,
            'colormap': self.colormap_combo.currentText(),
            'vmin': self.vmin_spin.value() if not self.auto_scale_check.isChecked() else None,
            'vmax': self.vmax_spin.value() if not self.auto_scale_check.isChecked() else None,
            'auto_scale': self.auto_scale_check.isChecked(),
            'tab_name': self.name_entry.text()
        }

        # Add range selection parameters
        range_selection = self.get_range_selection()
        params.update(range_selection)

        return params

    def get_range_selection(self):
        """Get current range selection settings."""
        range_type = self.range_button_group.checkedId()

        if range_type == 0:  # Full Range
            return {
                'range_type': 'full',
                'start_frame': None,
                'end_frame': None
            }
        elif range_type == 1:  # Time Range
            start_time = self.start_time_spin.value()
            end_time = self.end_time_spin.value()

            # Get dataset bounds for validation
            min_time, max_time = self.get_dataset_time_bounds()

            # Validate and clamp with user feedback
            start_time_clamped, end_time_clamped = self.validate_and_clamp_range(
                start_time, end_time, min_time, max_time, "time (minutes)"
            )

            # Convert to frames
            start_frame = self.time_to_frame(start_time_clamped)
            end_frame = self.time_to_frame(end_time_clamped)

            return {
                'range_type': 'time',
                'start_time': start_time_clamped,
                'end_time': end_time_clamped,
                'start_frame': start_frame,
                'end_frame': end_frame
            }
        elif range_type == 2:  # Frame Range
            start_frame = self.start_frame_spin.value()
            end_frame = self.end_frame_spin.value()

            # Get dataset bounds for validation
            min_frame, max_frame = self.get_dataset_frame_bounds()

            # Validate and clamp with user feedback
            start_frame_clamped, end_frame_clamped = self.validate_and_clamp_range(
                start_frame, end_frame, min_frame, max_frame, "frame"
            )

            return {
                'range_type': 'frame',
                'start_frame': start_frame_clamped,
                'end_frame': end_frame_clamped
            }
        else:
            # Default to full range
            return {
                'range_type': 'full',
                'start_frame': None,
                'end_frame': None
            }

    def time_to_frame(self, time_minutes):
        """Convert time in minutes to frame number."""
        # Using the existing conversion: frames = time_minutes * 60 / 0.02
        return int(time_minutes * 60 / 0.02)

    def frame_to_time(self, frame_number):
        """Convert frame number to time in minutes."""
        # Using the existing conversion: time = (frame * 0.02) / 60
        return (frame_number * 0.02) / 60

    def get_dataset_frame_bounds(self):
        """Get the actual min/max frame numbers from the dataset."""
        if not self.dataset:
            return 0, 1000  # Default fallback

        try:
            peak_idx = self.current_peak_idx
            frames = self.dataset.frame_numbers[peak_idx, :].compute()
            valid_frames = frames[frames != 0]

            if len(valid_frames) == 0:
                return 0, 1000  # Default fallback

            return int(valid_frames.min()), int(valid_frames.max())
        except Exception:
            return 0, 1000  # Default fallback

    def get_dataset_time_bounds(self):
        """Get the actual min/max time values from the dataset."""
        min_frame, max_frame = self.get_dataset_frame_bounds()
        min_time = self.frame_to_time(min_frame)
        max_time = self.frame_to_time(max_frame)
        return min_time, max_time

    def clamp_value(self, value, min_val, max_val):
        """Clamp a value to be within the specified bounds."""
        return max(min_val, min(value, max_val))

    def validate_and_clamp_range(self, start_val, end_val, min_val, max_val, value_type="value"):
        """Validate and clamp range values, providing user feedback for adjustments."""
        original_start, original_end = start_val, end_val

        # Clamp values to bounds
        start_clamped = self.clamp_value(start_val, min_val, max_val)
        end_clamped = self.clamp_value(end_val, min_val, max_val)

        # Ensure start <= end
        if start_clamped > end_clamped:
            start_clamped, end_clamped = end_clamped, start_clamped

        # Provide feedback if values were adjusted
        adjustments = []
        if original_start != start_clamped:
            if original_start < min_val:
                adjustments.append(f"Start {value_type} {original_start:.2f} was too low, adjusted to {start_clamped:.2f}")
            elif original_start > max_val:
                adjustments.append(f"Start {value_type} {original_start:.2f} was too high, adjusted to {start_clamped:.2f}")

        if original_end != end_clamped:
            if original_end < min_val:
                adjustments.append(f"End {value_type} {original_end:.2f} was too low, adjusted to {end_clamped:.2f}")
            elif original_end > max_val:
                adjustments.append(f"End {value_type} {original_end:.2f} was too high, adjusted to {end_clamped:.2f}")

        if original_start > original_end:
            adjustments.append(f"Start/end {value_type} values were swapped to maintain proper order")

        # Show warning if adjustments were made
        if adjustments:
            from PyQt5.QtWidgets import QMessageBox
            message = f"Range values were automatically adjusted:\n\n" + "\n".join(adjustments)
            message += f"\n\nDataset {value_type} range: {min_val:.2f} to {max_val:.2f}"
            QMessageBox.information(self, "Range Adjusted", message)

        return start_clamped, end_clamped

    def on_range_type_changed(self, button):
        """Handle range type radio button changes."""
        range_type = self.range_button_group.checkedId()

        if range_type == 0:  # Full Range
            self.start_time_spin.setEnabled(False)
            self.end_time_spin.setEnabled(False)
            self.start_frame_spin.setEnabled(False)
            self.end_frame_spin.setEnabled(False)
            self.auto_detect_btn.setEnabled(False)
        elif range_type == 1:  # Time Range
            self.start_time_spin.setEnabled(True)
            self.end_time_spin.setEnabled(True)
            self.start_frame_spin.setEnabled(False)
            self.end_frame_spin.setEnabled(False)
            self.auto_detect_btn.setEnabled(True)
        elif range_type == 2:  # Frame Range
            self.start_time_spin.setEnabled(False)
            self.end_time_spin.setEnabled(False)
            self.start_frame_spin.setEnabled(True)
            self.end_frame_spin.setEnabled(True)
            self.auto_detect_btn.setEnabled(True)

    def auto_detect_range(self):
        """Auto-detect and populate range values from dataset."""
        if not self.dataset:
            return

        try:
            # Get frame numbers for current peak
            peak_idx = self.current_peak_idx
            frames = self.dataset.frame_numbers[peak_idx, :].compute()

            # Remove zero frames (unfilled slots)
            valid_frames = frames[frames != 0]
            if len(valid_frames) == 0:
                return

            min_frame = int(valid_frames.min())
            max_frame = int(valid_frames.max())

            range_type = self.range_button_group.checkedId()

            if range_type == 1:  # Time Range
                min_time = self.frame_to_time(min_frame)
                max_time = self.frame_to_time(max_frame)
                self.start_time_spin.setValue(min_time)
                self.end_time_spin.setValue(max_time)
            elif range_type == 2:  # Frame Range
                self.start_frame_spin.setValue(min_frame)
                self.end_frame_spin.setValue(max_frame)

        except Exception as e:
            QMessageBox.warning(self, "Auto-detect Error", f"Could not auto-detect range: {str(e)}")

    def calculate_selection_average(self):
        """Calculate the average of the selected area."""
        if not self.dataset:
            self.selection_average_label.setText("Average of Selection: No dataset")
            return

        try:
            # Get selection parameters
            frame_start = self.selection_start_frame_spin.value()
            frame_end = self.selection_end_frame_spin.value()
            azimuth_start = self.selection_start_azimuth_spin.value()
            azimuth_end = self.selection_end_azimuth_spin.value()

            # Get current measurement and peak
            measurement = self.current_measurement
            peak_idx = self.current_peak_idx

            # Validate measurement exists
            if measurement not in self.dataset.col_idx:
                self.selection_average_label.setText("Average of Selection: Invalid measurement")
                return

            # Get data
            measurement_idx = self.dataset.col_idx[measurement]
            data_slice = self.dataset.data[peak_idx, :, :, measurement_idx].compute()
            frames = self.dataset.frame_numbers[peak_idx, :].compute()
            azimuths = self.dataset.azimuth_angles[peak_idx, :].compute()

            # Remove zero/invalid data
            valid_az_mask = azimuths != 0
            azimuths = azimuths[valid_az_mask]
            data_slice = data_slice[:, valid_az_mask]

            valid_frame_mask = frames != 0
            if len(frames) > 0 and frames[0] == 0:
                valid_frame_mask[0] = True
            frames = frames[valid_frame_mask]
            data_slice = data_slice[valid_frame_mask, :]

            # Create frame and azimuth masks for selection
            frame_mask = (frames >= frame_start) & (frames <= frame_end)
            azimuth_mask = (azimuths >= azimuth_start) & (azimuths <= azimuth_end)

            # Extract selected data
            selected_data = data_slice[frame_mask, :][:, azimuth_mask]

            if selected_data.size == 0:
                self.selection_average_label.setText("Average of Selection: No data in range")
                return

            # Calculate average
            avg_value = np.mean(selected_data[selected_data != 0])  # Exclude zeros

            # Format based on measurement type
            if 'strain' in measurement:
                formatted_avg = f"{avg_value:.6f}"
            elif 'd' in measurement:
                formatted_avg = f"{avg_value:.4f}"
            else:
                formatted_avg = f"{avg_value:.4f}"

            self.selection_average_label.setText(f"Average of Selection: {formatted_avg}")

            # Signal for drawing selection box (will implement next)
            self.emit_selection_box_signal(frame_start, frame_end, azimuth_start, azimuth_end)

        except Exception as e:
            self.selection_average_label.setText(f"Average of Selection: Error - {str(e)}")

    def emit_selection_box_signal(self, frame_start, frame_end, azimuth_start, azimuth_end):
        """Signal the main window to draw selection box on plot."""
        # Find the main DataAnalyzer window by traversing parents
        widget = self
        while widget is not None:
            if isinstance(widget, DataAnalyzer):
                widget.draw_selection_on_current_plot(frame_start, frame_end, azimuth_start, azimuth_end)
                return
            widget = widget.parent()
        print("Error: Could not find DataAnalyzer parent window")

    def update_selection_preview(self):
        """Update selection box preview when range inputs change."""
        try:
            frame_start = self.selection_start_frame_spin.value()
            frame_end = self.selection_end_frame_spin.value()
            azimuth_start = self.selection_start_azimuth_spin.value()
            azimuth_end = self.selection_end_azimuth_spin.value()

            # Find the main window and check if we have valid data
            main_window = None
            widget = self
            while widget is not None:
                if isinstance(widget, DataAnalyzer):
                    main_window = widget
                    break
                widget = widget.parent()

            # Only draw preview if we have valid data and a current plot
            if (hasattr(self, 'dataset') and self.dataset and
                main_window and hasattr(main_window, 'current_plot_data') and
                main_window.current_plot_data is not None):
                self.emit_selection_box_signal(frame_start, frame_end, azimuth_start, azimuth_end)

        except Exception as e:
            print(f"Error updating selection preview: {e}")


class DataAnalyzer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("XRD Data Analyzer v3.0 - Enhanced Multi-Analysis")

        # Initialize settings manager
        self.settings_manager = SettingsManager()
        self.current_settings = self.settings_manager.load_settings()

        # Apply window settings
        window_settings = self.current_settings["window"]
        self.setGeometry(
            window_settings["position"][0],
            window_settings["position"][1],
            window_settings["size"][0],
            window_settings["size"][1]
        )

        # Initialize variables
        self.current_dataset = None
        self.analysis_counter = 1
        self.images_gallery = []  # Store generated images
        self.gallery_view_mode = self.current_settings["ui_preferences"]["gallery_view_mode"]

        # Auto-save timer to debounce settings saves
        self.settings_save_timer = QtCore.QTimer()
        self.settings_save_timer.setSingleShot(True)
        self.settings_save_timer.timeout.connect(self.save_current_settings)

        self.setup_ui()

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Create toolbar
        toolbar = self.addToolBar("Main")
        toolbar.setMovable(False)

        # Add global restore defaults button
        restore_defaults_action = toolbar.addAction("🔄 Restore Defaults")
        restore_defaults_action.setToolTip("Reset all settings to factory defaults")
        restore_defaults_action.triggered.connect(self.restore_all_defaults)

        toolbar.addSeparator()

        # Main layout
        main_layout = QHBoxLayout()

        # Left panel - Dataset management and Gallery
        left_panel = QWidget()
        left_panel.setMaximumWidth(400)
        left_layout = QVBoxLayout()

        # Dataset loading group
        dataset_group = QGroupBox("Datasets")
        dataset_layout = QVBoxLayout()

        load_btn = QPushButton("Load Zarr Dataset")
        load_btn.clicked.connect(self.load_dataset)

        browse_btn = QPushButton("Browse Zarr Folder")
        browse_btn.clicked.connect(self.browse_zarr_folder)

        clear_btn = QPushButton("Clear All")
        clear_btn.clicked.connect(self.clear_datasets)

        self.dataset_list = DatasetListWidget()
        self.dataset_list.itemSelectionChanged.connect(self.on_dataset_selection_changed)

        dataset_layout.addWidget(load_btn)
        dataset_layout.addWidget(browse_btn)
        dataset_layout.addWidget(self.dataset_list)
        dataset_layout.addWidget(clear_btn)
        dataset_group.setLayout(dataset_layout)

        # Image Gallery placeholder
        gallery_group = QGroupBox("Image Gallery")
        gallery_layout = QVBoxLayout()

        # Gallery view buttons
        gallery_controls = QHBoxLayout()
        self.grid_view_btn = QPushButton("Grid View")
        self.grid_view_btn.setCheckable(True)
        self.grid_view_btn.setChecked(True)

        self.carousel_view_btn = QPushButton("Carousel View")
        self.carousel_view_btn.setCheckable(True)

        clear_gallery_btn = QPushButton("Clear Gallery")
        clear_gallery_btn.clicked.connect(self.clear_gallery)

        gallery_controls.addWidget(self.grid_view_btn)
        gallery_controls.addWidget(self.carousel_view_btn)
        gallery_controls.addWidget(clear_gallery_btn)

        # Gallery content area (placeholder for now)
        self.gallery_widget = QScrollArea()
        self.gallery_content = QWidget()
        self.gallery_layout = QGridLayout(self.gallery_content)
        self.gallery_widget.setWidget(self.gallery_content)
        self.gallery_widget.setWidgetResizable(True)
        self.gallery_widget.setMaximumHeight(200)

        gallery_layout.addLayout(gallery_controls)
        gallery_layout.addWidget(self.gallery_widget)
        gallery_group.setLayout(gallery_layout)

        # Batch Analysis group
        batch_group = QGroupBox("Batch Analysis")
        batch_layout = QFormLayout()

        # Frame range for batch processing
        self.batch_start_frame_spin = QSpinBox()
        self.batch_start_frame_spin.setRange(0, 999999)
        self.batch_start_frame_spin.setValue(self.current_settings["batch_analysis"]["frame_start"])

        self.batch_end_frame_spin = QSpinBox()
        self.batch_end_frame_spin.setRange(0, 999999)
        self.batch_end_frame_spin.setValue(self.current_settings["batch_analysis"]["frame_end"])

        # Azimuth ranges management
        ranges_controls_layout = QHBoxLayout()
        self.add_range_btn = QPushButton("Add Range")
        self.add_range_btn.clicked.connect(self.add_azimuth_range)
        self.edit_range_btn = QPushButton("Edit")
        self.edit_range_btn.clicked.connect(self.edit_azimuth_range)
        self.delete_range_btn = QPushButton("Delete")
        self.delete_range_btn.clicked.connect(self.delete_azimuth_range)
        self.reset_ranges_btn = QPushButton("Reset Defaults")
        self.reset_ranges_btn.clicked.connect(self.reset_default_ranges)

        ranges_controls_layout.addWidget(self.add_range_btn)
        ranges_controls_layout.addWidget(self.edit_range_btn)
        ranges_controls_layout.addWidget(self.delete_range_btn)
        ranges_controls_layout.addWidget(self.reset_ranges_btn)

        # Azimuth ranges list widget
        self.azimuth_ranges_list = QListWidget()
        self.azimuth_ranges_list.setMaximumHeight(110)  # Slightly increased for larger text
        self.azimuth_ranges_list.setStyleSheet("font-size: 12px;")

        # Initialize azimuth ranges from settings
        self.azimuth_range_data = self.current_settings["batch_analysis"]["azimuth_ranges"].copy()
        self.update_ranges_display()

        # Run batch analysis button
        self.run_batch_btn = QPushButton("Run Batch Analysis")
        self.run_batch_btn.clicked.connect(self.run_batch_analysis)
        self.run_batch_btn.setStyleSheet("QPushButton { background-color: #e8f4f8; font-weight: bold; padding: 8px; }")

        # Progress bar
        self.batch_progress = QProgressBar()
        self.batch_progress.setVisible(False)

        # Results summary
        self.batch_results_label = QLabel("Results: --")
        self.batch_results_label.setStyleSheet("font-size: 10px; color: #333; padding: 2px; background-color: #f9f9f9; border: 1px solid #ddd;")

        # Export buttons
        export_layout = QHBoxLayout()
        self.export_csv_btn = QPushButton("Export CSV")
        self.export_csv_btn.clicked.connect(self.export_batch_csv)
        self.export_csv_btn.setEnabled(False)

        self.export_excel_btn = QPushButton("Export Excel")
        self.export_excel_btn.clicked.connect(self.export_batch_excel)
        self.export_excel_btn.setEnabled(False)

        export_layout.addWidget(self.export_csv_btn)
        export_layout.addWidget(self.export_excel_btn)

        # Add to layout
        batch_layout.addRow("Frame Start:", self.batch_start_frame_spin)
        batch_layout.addRow("Frame End:", self.batch_end_frame_spin)
        batch_layout.addRow("Range Controls:", ranges_controls_layout)
        batch_layout.addRow("Azimuth Ranges:", self.azimuth_ranges_list)
        batch_layout.addRow("", self.run_batch_btn)
        batch_layout.addRow("", self.batch_progress)
        batch_layout.addRow("Results:", self.batch_results_label)
        batch_layout.addRow("Export:", export_layout)

        batch_group.setLayout(batch_layout)

        left_layout.addWidget(dataset_group)
        left_layout.addWidget(gallery_group)
        left_layout.addWidget(batch_group)
        left_layout.addStretch()
        left_panel.setLayout(left_layout)

        # Right panel - Tabbed Analysis and Plot area
        right_panel = QWidget()
        right_layout = QVBoxLayout()

        # Analysis tabs with controls
        tabs_and_controls_layout = QHBoxLayout()

        # Analysis tabs widget
        self.analysis_tabs = QTabWidget()
        self.analysis_tabs.setTabsClosable(True)
        self.analysis_tabs.tabCloseRequested.connect(self.close_analysis_tab)

        # Tab control buttons
        tab_controls_layout = QVBoxLayout()

        self.add_tab_btn = QPushButton("+ Add Analysis")
        self.add_tab_btn.clicked.connect(self.add_analysis_tab)
        self.add_tab_btn.setStyleSheet("QPushButton { font-weight: bold; background-color: #e8f4f8; }")

        self.duplicate_tab_btn = QPushButton("Duplicate Tab")
        self.duplicate_tab_btn.clicked.connect(self.duplicate_current_tab)

        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(self.export_csv)

        save_plot_btn = QPushButton("Save Plot")
        save_plot_btn.clicked.connect(self.save_plot)

        # Generate plot button - moved from inside tabs to here, made bigger
        generate_plot_btn = QPushButton("Generate Plot")
        generate_plot_btn.clicked.connect(self.generate_current_tab_plot)
        generate_plot_btn.setStyleSheet("QPushButton { font-weight: bold; padding: 12px; background-color: #4CAF50; color: white; font-size: 14px; }")
        generate_plot_btn.setMinimumHeight(50)

        tab_controls_layout.addWidget(self.add_tab_btn)
        tab_controls_layout.addWidget(self.duplicate_tab_btn)
        tab_controls_layout.addWidget(export_csv_btn)
        tab_controls_layout.addWidget(save_plot_btn)
        tab_controls_layout.addWidget(generate_plot_btn)
        tab_controls_layout.addStretch()

        tab_controls_widget = QWidget()
        tab_controls_widget.setMaximumWidth(180)  # Increased width for larger generate button
        tab_controls_widget.setLayout(tab_controls_layout)

        tabs_and_controls_layout.addWidget(self.analysis_tabs)
        tabs_and_controls_layout.addWidget(tab_controls_widget)

        # Info label and plot area
        plot_area_layout = QVBoxLayout()

        self.info_label = QLabel("Load a dataset and add an analysis to begin")
        self.info_label.setAlignment(Qt.AlignCenter)
        self.info_label.setStyleSheet("color: gray; font-size: 14px;")

        # Enhanced plot canvas
        self.plot_canvas = PlotCanvas()

        # Store current plot data for selection box drawing
        self.current_plot_data = None

        # Store batch analysis results
        self.batch_results = []

        plot_area_layout.addWidget(self.info_label)
        plot_area_layout.addWidget(self.plot_canvas)

        # Add to right panel with fixed height for tabs and maximum space for plot
        tabs_and_controls_widget = QWidget()
        tabs_and_controls_widget.setMaximumHeight(400)  # Fixed height for analysis tabs area
        tabs_and_controls_widget.setLayout(tabs_and_controls_layout)

        right_layout.addWidget(tabs_and_controls_widget)    # Fixed height
        right_layout.addLayout(plot_area_layout, 1)         # All remaining space for plot

        right_panel.setLayout(right_layout)

        # Add panels to main layout
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([400, 1000])

        main_layout.addWidget(splitter)
        central_widget.setLayout(main_layout)

        # Create initial analysis tab
        self.add_analysis_tab()

        # Initialize gallery display
        self.set_grid_view()  # This will also call update_gallery_display()

        # Apply saved settings to UI and connect auto-save signals
        self.apply_settings_to_ui()
        self.connect_ui_signals_for_auto_save()

        # Update UI state
        self.update_ui_state()

    def add_analysis_tab(self, tab_name=None):
        """Add a new analysis tab."""
        if tab_name is None:
            tab_count = self.analysis_tabs.count()
            tab_name = f"Analysis {tab_count + 1}"

        analysis_tab = AnalysisTab(tab_name, self.current_dataset)
        analysis_tab.plotRequested.connect(self.on_plot_requested)

        index = self.analysis_tabs.addTab(analysis_tab, tab_name)
        self.analysis_tabs.setCurrentIndex(index)

        # Tab close functionality is handled by built-in tab close buttons

    def close_analysis_tab(self, index=None):
        """Close the analysis tab at the given index (or current tab if None)."""
        if self.analysis_tabs.count() <= 1:
            return

        if index is None:
            index = self.analysis_tabs.currentIndex()

        self.analysis_tabs.removeTab(index)

    def duplicate_current_tab(self):
        """Duplicate the current analysis tab with same settings."""
        current_tab = self.analysis_tabs.currentWidget()
        if current_tab is None:
            return

        # Create new tab with copied settings
        tab_count = self.analysis_tabs.count()
        new_tab_name = f"Analysis {tab_count + 1}"

        new_tab = AnalysisTab(new_tab_name, self.current_dataset)
        new_tab.plotRequested.connect(self.on_plot_requested)

        # Copy settings from current tab
        new_tab.measurement_combo.setCurrentText(current_tab.measurement_combo.currentText())
        new_tab.delta_check.setChecked(current_tab.delta_check.isChecked())
        new_tab.abs_check.setChecked(current_tab.abs_check.isChecked())
        new_tab.pct_check.setChecked(current_tab.pct_check.isChecked())
        new_tab.peak_combo.setCurrentText(current_tab.peak_combo.currentText())
        new_tab.colormap_combo.setCurrentText(current_tab.colormap_combo.currentText())
        new_tab.graphing_mode_combo.setCurrentText(current_tab.graphing_mode_combo.currentText())
        new_tab.auto_scale_check.setChecked(current_tab.auto_scale_check.isChecked())
        new_tab.vmin_spin.setValue(current_tab.vmin_spin.value())
        new_tab.vmax_spin.setValue(current_tab.vmax_spin.value())

        # Copy range settings
        current_range_type = current_tab.range_button_group.checkedId()
        if current_range_type == 0:
            new_tab.full_range_radio.setChecked(True)
        elif current_range_type == 1:
            new_tab.time_range_radio.setChecked(True)
        elif current_range_type == 2:
            new_tab.frame_range_radio.setChecked(True)

        new_tab.start_time_spin.setValue(current_tab.start_time_spin.value())
        new_tab.end_time_spin.setValue(current_tab.end_time_spin.value())
        new_tab.start_frame_spin.setValue(current_tab.start_frame_spin.value())
        new_tab.end_frame_spin.setValue(current_tab.end_frame_spin.value())

        # Trigger the range type change to update enabled state
        new_tab.on_range_type_changed(None)

        # Update the new tab's current_measurement based on copied settings
        new_tab.update_current_measurement()

        index = self.analysis_tabs.addTab(new_tab, new_tab_name)
        self.analysis_tabs.setCurrentIndex(index)

    def on_plot_requested(self, params, tab_name):
        """Handle plot request from analysis tab."""
        self.plot_heatmap(params, tab_name)

    def generate_current_tab_plot(self):
        """Generate plot from the currently active analysis tab."""
        current_tab = self.analysis_tabs.currentWidget()
        if current_tab and hasattr(current_tab, 'request_generation'):
            # Update batch frame defaults if this is a frame-ranged plot
            self.update_batch_frame_defaults_from_current_tab()
            current_tab.request_generation()
        else:
            QMessageBox.warning(self, "No Active Tab", "Please select an analysis tab to generate a plot.")

    def update_batch_frame_defaults_from_current_tab(self):
        """Update batch analysis frame ranges based on current tab's frame range settings."""
        current_tab = self.analysis_tabs.currentWidget()
        if not current_tab or not hasattr(current_tab, 'get_range_selection'):
            return

        try:
            # Get the current tab's range selection
            range_selection = current_tab.get_range_selection()

            # Only update if this is a frame-based range
            if range_selection.get('range_type') == 'frame':
                end_frame = range_selection.get('end_frame')
                if end_frame is not None:
                    # Set batch end frame to match analysis tab's end frame
                    self.batch_end_frame_spin.setValue(end_frame)

                    # Set batch start frame to end frame minus 10 (minimum 0)
                    start_frame = max(0, end_frame - 10)
                    self.batch_start_frame_spin.setValue(start_frame)

                    print(f"Updated batch frame range: {start_frame} to {end_frame} (based on analysis tab frame range)")

        except Exception as e:
            print(f"Error updating batch frame defaults: {e}")

    def update_ui_state(self):
        """Update UI elements based on current state."""
        has_dataset = self.current_dataset is not None

        # Enable/disable tab controls
        self.add_tab_btn.setEnabled(has_dataset)
        self.duplicate_tab_btn.setEnabled(has_dataset)
        # Tab closing is handled by built-in tab close buttons

        # Update analysis tabs
        for i in range(self.analysis_tabs.count()):
            analysis_tab = self.analysis_tabs.widget(i)
            if analysis_tab:
                analysis_tab.update_for_dataset(self.current_dataset)

        # Update info label
        if has_dataset:
            info_text = f"Dataset: {self.current_dataset.params.sample} - {self.current_dataset.params.stage.name}\n"
            info_text += f"Shape: {self.current_dataset.data.shape}\n"
            if hasattr(self.current_dataset.params, 'active_peaks'):
                peak_names = [peak.name for peak in self.current_dataset.params.active_peaks]
                info_text += f"Peaks: {', '.join(peak_names)}"
            self.info_label.setText(info_text)
        else:
            self.info_label.setText("Load a dataset to begin analysis")

    def apply_settings_to_ui(self):
        """Apply loaded settings to all UI controls."""
        try:
            # Apply analysis defaults to all analysis tabs
            for i in range(self.analysis_tabs.count()):
                analysis_tab = self.analysis_tabs.widget(i)
                if analysis_tab:
                    self.apply_settings_to_analysis_tab(analysis_tab)
        except Exception as e:
            print(f"Error applying settings to UI: {e}")

    def apply_settings_to_analysis_tab(self, tab):
        """Apply settings to a specific analysis tab."""
        try:
            defaults = self.current_settings["analysis_defaults"]

            # Set colormap
            if hasattr(tab, 'colormap_combo'):
                index = tab.colormap_combo.findText(defaults["colormap"])
                if index >= 0:
                    tab.colormap_combo.setCurrentIndex(index)

            # Set scale values
            if hasattr(tab, 'vmin_spin'):
                tab.vmin_spin.setValue(defaults["vmin"])
            if hasattr(tab, 'vmax_spin'):
                tab.vmax_spin.setValue(defaults["vmax"])
            if hasattr(tab, 'auto_scale_check'):
                tab.auto_scale_check.setChecked(defaults["auto_scale"])

            # Set checkboxes
            if hasattr(tab, 'delta_check'):
                tab.delta_check.setChecked(defaults["delta_enabled"])
            if hasattr(tab, 'abs_check'):
                tab.abs_check.setChecked(defaults["abs_enabled"])
            if hasattr(tab, 'pct_check'):
                tab.pct_check.setChecked(defaults["pct_enabled"])

            # Set time/frame ranges
            if hasattr(tab, 'start_time_spin'):
                tab.start_time_spin.setValue(defaults["start_time"])
            if hasattr(tab, 'end_time_spin'):
                tab.end_time_spin.setValue(defaults["end_time"])
            if hasattr(tab, 'start_frame_spin'):
                tab.start_frame_spin.setValue(defaults["start_frame"])
            if hasattr(tab, 'end_frame_spin'):
                tab.end_frame_spin.setValue(defaults["end_frame"])

        except Exception as e:
            print(f"Error applying settings to analysis tab: {e}")

    def connect_ui_signals_for_auto_save(self):
        """Connect UI control signals to trigger auto-save."""
        try:
            # Connect batch analysis controls
            self.batch_start_frame_spin.valueChanged.connect(self.schedule_settings_save)
            self.batch_end_frame_spin.valueChanged.connect(self.schedule_settings_save)

            # Connect analysis tab signals for all tabs
            for i in range(self.analysis_tabs.count()):
                analysis_tab = self.analysis_tabs.widget(i)
                if analysis_tab:
                    self.connect_analysis_tab_signals(analysis_tab)
        except Exception as e:
            print(f"Error connecting UI signals: {e}")

    def connect_analysis_tab_signals(self, tab):
        """Connect signals for a specific analysis tab."""
        try:
            # Connect colormap changes
            if hasattr(tab, 'colormap_combo'):
                tab.colormap_combo.currentTextChanged.connect(self.schedule_settings_save)

            # Connect scale value changes
            if hasattr(tab, 'vmin_spin'):
                tab.vmin_spin.valueChanged.connect(self.schedule_settings_save)
            if hasattr(tab, 'vmax_spin'):
                tab.vmax_spin.valueChanged.connect(self.schedule_settings_save)
            if hasattr(tab, 'auto_scale_check'):
                tab.auto_scale_check.toggled.connect(self.schedule_settings_save)

            # Connect checkbox changes
            if hasattr(tab, 'delta_check'):
                tab.delta_check.toggled.connect(self.schedule_settings_save)
            if hasattr(tab, 'abs_check'):
                tab.abs_check.toggled.connect(self.schedule_settings_save)
            if hasattr(tab, 'pct_check'):
                tab.pct_check.toggled.connect(self.schedule_settings_save)

            # Connect time/frame range changes
            if hasattr(tab, 'start_time_spin'):
                tab.start_time_spin.valueChanged.connect(self.schedule_settings_save)
            if hasattr(tab, 'end_time_spin'):
                tab.end_time_spin.valueChanged.connect(self.schedule_settings_save)
            if hasattr(tab, 'start_frame_spin'):
                tab.start_frame_spin.valueChanged.connect(self.schedule_settings_save)
            if hasattr(tab, 'end_frame_spin'):
                tab.end_frame_spin.valueChanged.connect(self.schedule_settings_save)

        except Exception as e:
            print(f"Error connecting analysis tab signals: {e}")

    def schedule_settings_save(self):
        """Schedule a debounced settings save."""
        self.settings_save_timer.stop()
        self.settings_save_timer.start(500)  # 500ms delay

    def save_current_settings(self):
        """Save current UI state to settings."""
        try:
            # Update window settings
            self.current_settings["window"]["size"] = [self.width(), self.height()]
            self.current_settings["window"]["position"] = [self.x(), self.y()]

            # Update batch analysis settings
            self.current_settings["batch_analysis"]["frame_start"] = self.batch_start_frame_spin.value()
            self.current_settings["batch_analysis"]["frame_end"] = self.batch_end_frame_spin.value()
            self.current_settings["batch_analysis"]["azimuth_ranges"] = self.azimuth_range_data.copy()

            # Update UI preferences
            self.current_settings["ui_preferences"]["gallery_view_mode"] = self.gallery_view_mode

            # Update analysis defaults from first tab if available
            if self.analysis_tabs.count() > 0:
                first_tab = self.analysis_tabs.widget(0)
                if first_tab:
                    self.save_analysis_tab_settings(first_tab)

            # Save to file
            self.settings_manager.save_settings(self.current_settings)

        except Exception as e:
            print(f"Error saving settings: {e}")

    def save_analysis_tab_settings(self, tab):
        """Save settings from an analysis tab."""
        try:
            defaults = self.current_settings["analysis_defaults"]

            # Save colormap
            if hasattr(tab, 'colormap_combo'):
                defaults["colormap"] = tab.colormap_combo.currentText()

            # Save scale values
            if hasattr(tab, 'vmin_spin'):
                defaults["vmin"] = tab.vmin_spin.value()
            if hasattr(tab, 'vmax_spin'):
                defaults["vmax"] = tab.vmax_spin.value()
            if hasattr(tab, 'auto_scale_check'):
                defaults["auto_scale"] = tab.auto_scale_check.isChecked()

            # Save checkboxes
            if hasattr(tab, 'delta_check'):
                defaults["delta_enabled"] = tab.delta_check.isChecked()
            if hasattr(tab, 'abs_check'):
                defaults["abs_enabled"] = tab.abs_check.isChecked()
            if hasattr(tab, 'pct_check'):
                defaults["pct_enabled"] = tab.pct_check.isChecked()

            # Save time/frame ranges
            if hasattr(tab, 'start_time_spin'):
                defaults["start_time"] = tab.start_time_spin.value()
            if hasattr(tab, 'end_time_spin'):
                defaults["end_time"] = tab.end_time_spin.value()
            if hasattr(tab, 'start_frame_spin'):
                defaults["start_frame"] = tab.start_frame_spin.value()
            if hasattr(tab, 'end_frame_spin'):
                defaults["end_frame"] = tab.end_frame_spin.value()

        except Exception as e:
            print(f"Error saving analysis tab settings: {e}")

    def restore_all_defaults(self):
        """Restore all settings to factory defaults with confirmation."""
        reply = QMessageBox.question(
            self,
            "Restore Defaults",
            "Reset ALL settings to factory defaults?\n\n"
            "This will reset:\n"
            "• Analysis parameters (colors, scales, checkboxes)\n"
            "• Batch analysis settings\n"
            "• Azimuth ranges\n"
            "• Window size and position\n"
            "• All preferences\n\n"
            "Are you sure you want to continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # Reset to default settings
                self.current_settings = self.settings_manager.reset_to_defaults()

                # Apply window settings
                window_settings = self.current_settings["window"]
                self.setGeometry(
                    window_settings["position"][0],
                    window_settings["position"][1],
                    window_settings["size"][0],
                    window_settings["size"][1]
                )

                # Reset azimuth ranges
                self.azimuth_range_data = self.current_settings["batch_analysis"]["azimuth_ranges"].copy()
                self.update_ranges_display()

                # Reset batch frame values
                self.batch_start_frame_spin.setValue(self.current_settings["batch_analysis"]["frame_start"])
                self.batch_end_frame_spin.setValue(self.current_settings["batch_analysis"]["frame_end"])

                # Reset gallery view mode
                self.gallery_view_mode = self.current_settings["ui_preferences"]["gallery_view_mode"]

                # Apply settings to all analysis tabs
                self.apply_settings_to_ui()

                # Save the reset settings
                self.settings_manager.save_settings(self.current_settings)

                QMessageBox.information(self, "Settings Reset", "All settings have been restored to factory defaults.")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to restore defaults: {e}")

    def closeEvent(self, event):
        """Handle window close event to save settings."""
        self.save_current_settings()
        event.accept()

    def load_dataset(self):
        """Load a single Zarr dataset."""
        zarr_dir = QFileDialog.getExistingDirectory(
            self, "Select Zarr Dataset Directory"
        )

        if not zarr_dir:
            return

        try:
            # Try to load the dataset
            # This is a simplified load - in real implementation, you'd need
            # to recreate GSASParams or store them in metadata
            metadata_path = os.path.join(zarr_dir, "metadata.json")
            if not os.path.exists(metadata_path):
                QMessageBox.warning(self, "Error", "No metadata.json found in dataset")
                return

            with open(metadata_path, 'r') as f:
                metadata = json.load(f)

            # Create params from loaded metadata
            from XRD.core.gsas_processing import GSASParams, Stages, PeakParams

            # Load parameters from metadata
            params_data = metadata.get('params', {})

            # Reconstruct active_peaks from metadata
            active_peaks = []
            if 'active_peaks' in params_data and params_data['active_peaks']:
                print(f"Loading peaks from metadata:")

                # IMPORTANT: Reverse the peak order to match actual data structure
                # Load peaks in original order (no reversal needed)
                peaks_list = params_data['active_peaks']

                for i, peak_data in enumerate(peaks_list):
                    peak = PeakParams(
                        name=peak_data.get('name', 'Unknown Peak'),
                        miller_index=peak_data.get('miller_index', '211'),
                        position=peak_data.get('position', 8.46),
                        limits=tuple(peak_data.get('limits', [8.2, 8.8]))
                    )
                    print(f"  Index {i}: {peak.name} ({peak.miller_index}) at {peak.position}")
                    active_peaks.append(peak)
            else:
                # Fallback: detect peaks from dataset shape - use only generic names
                n_peaks = metadata['n_peaks'] if 'n_peaks' in metadata else 3
                active_peaks = []

                # Just create generic peaks without assuming anything about positions or ranges
                for i in range(n_peaks):
                    peak = PeakParams(f"Unknown {i+1}", f"hkl{i+1}", 8.0, (7.5, 8.5))
                    active_peaks.append(peak)

                print(f"Generated {len(active_peaks)} fallback peaks for dataset with {n_peaks} peaks")

            # Get stage from metadata
            stage_name = params_data.get('stage', 'CONT')
            stage = getattr(Stages, stage_name, Stages.CONT)

            loaded_params = GSASParams(
                home_dir="",
                images_path="",
                refs_path=None,
                control_file="",
                mask_file="",
                intplot_export=False,
                sample=params_data.get('sample', 'Unknown'),
                setting=params_data.get('setting', 'Unknown'),
                stage=stage,
                notes="",
                exposure=params_data.get('exposure', '019'),
                active_peaks=active_peaks,
                azimuths=tuple(params_data.get('azimuths', (-110, 110))),
                frames=tuple(params_data.get('frames', (0, 100))),
                spacing=params_data.get('spacing', 5),
                step=1,
                pixel_size=(172.0, 172.0),
                wavelength=0.1726,
                detector_size=(1475, 1679)
            )

            dataset = XRDDataset.load(zarr_dir, loaded_params)

            # Validate dataset structure
            if dataset.data is None:
                raise ValueError("Dataset has no data - possibly corrupted or incomplete processing")

            data_shape = dataset.data.shape
            if len(data_shape) != 4:
                raise ValueError(f"Invalid dataset structure: expected 4D data, got {len(data_shape)}D")

            n_peaks, n_frames, n_azimuths, n_measurements = data_shape
            if n_peaks == 0:
                raise ValueError("Dataset has no peaks - invalid data structure")
            if n_frames == 0:
                raise ValueError("Dataset has no frames - invalid data structure")
            if n_azimuths == 0:
                raise ValueError("Dataset has no azimuthal data - invalid data structure")
            if n_measurements == 0:
                raise ValueError("Dataset has no measurements - invalid data structure")

            # Validate measurements
            if not dataset.col_idx or not dataset.measurement_cols:
                raise ValueError("Dataset has no measurement columns - corrupted metadata")

            print(f"Dataset validation passed: {n_peaks} peaks, {n_frames} frames, {n_azimuths} azimuths, {n_measurements} measurements")

            self.dataset_list.add_dataset(zarr_dir, dataset)

            # Update all analysis tabs with the new dataset
            self.update_all_tabs_with_dataset(dataset)

            QMessageBox.information(self, "Success", f"Loaded dataset: {os.path.basename(zarr_dir)}")

        except Exception as e:
            error_msg = f"Failed to load dataset: {str(e)}"
            if "No such file or directory" in str(e):
                error_msg += "\n\nCheck that the Zarr dataset is complete and not corrupted."
            elif "Invalid dataset structure" in str(e):
                error_msg += "\n\nThis may be due to incomplete processing or incompatible dataset version."
            elif "corrupted metadata" in str(e):
                error_msg += "\n\nTry regenerating the dataset using the batch processor."

            QMessageBox.critical(self, "Dataset Load Error", error_msg)

    def browse_zarr_folder(self):
        """Browse and load multiple Zarr datasets from a folder with progress dialog."""
        base_dir = QFileDialog.getExistingDirectory(
            self, "Select Folder Containing Zarr Datasets"
        )

        if not base_dir:
            return

        # Save last used folder
        self.current_settings["paths"]["last_zarr_folder"] = base_dir
        self.schedule_settings_save()

        # Find all zarr directories
        zarr_dirs = []
        for item in os.listdir(base_dir):
            item_path = os.path.join(base_dir, item)
            if os.path.isdir(item_path):
                # Check if it's a zarr dataset (has metadata.json)
                if os.path.exists(os.path.join(item_path, "metadata.json")):
                    zarr_dirs.append(item_path)

        if not zarr_dirs:
            QMessageBox.information(self, "No Datasets", "No Zarr datasets found in the selected folder")
            return

        # Show progress dialog and start loading
        progress_dialog = BulkLoadProgressDialog(self)

        # Create worker thread
        self.load_worker = BulkLoadWorker(zarr_dirs)

        # Connect worker signals to dialog
        self.load_worker.progress_updated.connect(
            lambda current, dataset_name: progress_dialog.update_progress(current, len(zarr_dirs), dataset_name)
        )
        self.load_worker.dataset_loaded.connect(self.on_dataset_loaded)
        self.load_worker.load_error.connect(progress_dialog.add_error)
        self.load_worker.finished.connect(progress_dialog.loading_finished)

        # Connect dialog cancel to worker stop
        progress_dialog.rejected.connect(self.load_worker.stop)

        # Start loading
        self.load_worker.start()

        # Show dialog (blocks until loading complete or cancelled)
        progress_dialog.exec_()

        # Clean up worker
        if self.load_worker.isRunning():
            self.load_worker.stop()
            self.load_worker.wait()

    def on_dataset_loaded(self, dataset_path, dataset):
        """Handle a successfully loaded dataset."""
        try:
            # Add to dataset list
            self.dataset_list.add_dataset(dataset_path, dataset)

            # Update analysis tabs with the first dataset loaded
            if self.dataset_list.count() == 1:
                self.update_all_tabs_with_dataset(dataset)

        except Exception as e:
            print(f"Error handling loaded dataset {dataset_path}: {e}")

    def clear_datasets(self):
        """Clear all loaded datasets."""
        self.dataset_list.clear_datasets()
        self.current_dataset = None
        self.plot_canvas.clear_plot()
        self.update_ui_state()

    def on_dataset_selection_changed(self):
        """Handle dataset selection change."""
        selected = self.dataset_list.get_selected_datasets()
        if selected:
            self.current_dataset = selected[0]  # Use first selected
        else:
            self.current_dataset = None

        self.update_ui_state()

    # Note: Peak combo, measurement selection, and other analysis controls
    # are now handled individually by each AnalysisTab

    # Note: Plot creation is now handled by the plot_heatmap method
    # which receives parameters from individual AnalysisTab instances

    def plot_heatmap(self, params, mode="generate"):
        """Create heatmap plot from analysis tab parameters."""
        if not self.current_dataset:
            QMessageBox.warning(self, "No Dataset", "Please select a dataset first")
            return

        try:
            # Extract measurement and handle delta/abs/pct variants
            measurement = params['measurement']
            is_delta = measurement.startswith('delta ')
            is_abs = measurement.startswith('abs ')
            is_pct = measurement.startswith('pct ')

            # Check if measurement exists in dataset
            if measurement in self.current_dataset.col_idx:
                measurement_idx = self.current_dataset.col_idx[measurement]
            else:
                # For delta/abs measurements, try to calculate them if base measurement exists
                if is_delta:
                    base_measurement = measurement.replace('delta ', '')
                    if base_measurement in self.current_dataset.col_idx:
                        # Calculate delta if it doesn't exist
                        self.current_dataset.calculate_delta(base_measurement)
                        measurement_idx = self.current_dataset.col_idx[measurement]
                    else:
                        QMessageBox.warning(self, "Invalid Measurement",
                                          f"Base measurement '{base_measurement}' not found for delta calculation")
                        return
                elif is_abs:
                    # For abs measurements, look for both "abs measurement" format and base measurement
                    base_measurement = measurement.replace('abs ', '')
                    if f"abs {base_measurement}" in self.current_dataset.col_idx:
                        measurement_idx = self.current_dataset.col_idx[f"abs {base_measurement}"]
                    elif base_measurement in self.current_dataset.col_idx:
                        # Calculate abs using add_measurement with numpy abs
                        base_idx = self.current_dataset.col_idx[base_measurement]
                        base_data = self.current_dataset.data[:, :, :, base_idx]
                        if hasattr(base_data, 'compute'):
                            base_data = base_data.compute()
                        abs_data = np.abs(base_data)
                        self.current_dataset.add_measurement(f"abs {base_measurement}", abs_data)
                        measurement_idx = self.current_dataset.col_idx[f"abs {base_measurement}"]
                    else:
                        QMessageBox.warning(self, "Invalid Measurement",
                                          f"Base measurement '{base_measurement}' not found for abs calculation")
                        return
                elif is_pct:
                    # For pct measurements, calculate them if base measurement exists
                    base_measurement = measurement.replace('pct ', '')
                    if base_measurement in self.current_dataset.col_idx:
                        # Calculate pct if it doesn't exist
                        self.current_dataset.calculate_pct(base_measurement)
                        measurement_idx = self.current_dataset.col_idx[measurement]
                    else:
                        QMessageBox.warning(self, "Invalid Measurement",
                                          f"Base measurement '{base_measurement}' not found for pct calculation")
                        return
                else:
                    QMessageBox.warning(self, "Invalid Measurement",
                                      f"Measurement '{measurement}' not found in dataset")
                    return

            # Extract 2D data (peak, frame, azimuth, measurement)
            peak_idx = params['peak_idx']
            data_slice = self.current_dataset.data[peak_idx, :, :, measurement_idx]

            # Convert to numpy array if it's a Dask array
            if hasattr(data_slice, 'compute'):
                data_slice = data_slice.compute()

            # Get frame and azimuth arrays for proper indexing
            frames = self.current_dataset.frame_numbers[peak_idx, :].compute()
            azimuths = self.current_dataset.azimuth_angles[peak_idx, :].compute()

            # Remove any zero azimuths (unfilled slots)
            valid_az_mask = azimuths != 0
            azimuths = azimuths[valid_az_mask]
            data_slice = data_slice[:, valid_az_mask]

            # Remove any zero frames (unfilled slots)
            valid_frame_mask = frames != 0
            # Keep first frame even if it's 0 (frame 0 is valid)
            if len(frames) > 0 and frames[0] == 0:
                valid_frame_mask[0] = True
            frames = frames[valid_frame_mask]
            data_slice = data_slice[valid_frame_mask, :]

            # Apply range filtering if specified
            if params.get('range_type') != 'full' and params.get('start_frame') is not None and params.get('end_frame') is not None:
                start_frame = params['start_frame']
                end_frame = params['end_frame']

                # Create range mask
                range_mask = (frames >= start_frame) & (frames <= end_frame)

                # Apply range mask
                frames = frames[range_mask]
                data_slice = data_slice[range_mask, :]

            # Convert frame numbers based on stage type (in-situ vs ex-situ)
            is_in_situ = self.current_dataset.params.stage.name == 'CONT'
            if is_in_situ:
                # Convert to time in minutes for in-situ (continuous) data
                time_values = (frames * 0.02) / 60
                df_index = time_values
                ylabel = "Time (min)"
            else:
                # Convert to depth for ex-situ data (BEF/AFT) - surface at bottom
                # Create depth values starting from 0 (surface) at bottom of plot
                depth_values = np.arange(len(frames))
                df_index = depth_values
                ylabel = "Depth (μm)"

            # Create DataFrame with proper orientation:
            # Rows = time/depth (y-axis), Columns = azimuths (x-axis)
            df = pd.DataFrame(data_slice, index=df_index, columns=azimuths)

            # Remove columns (azimuths) that are all NaN or zero
            df = df.loc[:, (df != 0).any(axis=0)]

            # Remove rows (frames) that are all NaN or zero
            df = df.loc[(df != 0).any(axis=1)]

            # Sort by index to ensure proper ordering
            df = df.sort_index()

            # Sort columns (azimuths) to ensure proper ordering
            df = df.reindex(sorted(df.columns), axis=1)

            # Apply graphing mode adjustments
            graphing_mode = params.get('mode', 'Standard')
            if graphing_mode == 'Robust':
                # Use robust scaling (percentile-based)
                if params['auto_scale']:
                    vmin = np.percentile(df.values, 5)
                    vmax = np.percentile(df.values, 95)
                else:
                    vmin, vmax = params['vmin'], params['vmax']
            elif 'L.L.' in graphing_mode:
                # Locked Limits mode
                vmin, vmax = params['vmin'], params['vmax']
            else:
                # Standard mode
                if params['auto_scale']:
                    vmin, vmax = None, None
                else:
                    vmin, vmax = params['vmin'], params['vmax']

            # Generate colorbar label based on measurement type
            if measurement.startswith("delta"):
                base_name = measurement.split()[-1]
                if base_name == 'd':
                    cbar_label = "Δ d-spacing (Å)"
                elif base_name == 'gamma':
                    cbar_label = "Δ FWHM (centidegrees)"
                elif base_name == 'sigma':
                    cbar_label = "Δ Peak Width"
                elif base_name == 'area':
                    cbar_label = "Δ Peak Area"
                elif base_name == 'pos':
                    cbar_label = "Δ Peak Position"
                else:
                    cbar_label = f"Δ {base_name.capitalize()}"
            elif measurement.startswith("abs"):
                base_name = measurement.split()[-1]
                if base_name == 'd':
                    cbar_label = "Abs d-spacing (Å)"
                elif base_name == 'gamma':
                    cbar_label = "Abs FWHM (centidegrees)"
                elif base_name == 'sigma':
                    cbar_label = "Abs Peak Width"
                elif base_name == 'area':
                    cbar_label = "Abs Peak Area"
                elif base_name == 'pos':
                    cbar_label = "Abs Peak Position"
                else:
                    cbar_label = f"Abs {base_name.capitalize()}"
            elif measurement.startswith("pct"):
                base_name = measurement.split()[-1]
                if base_name == 'd':
                    cbar_label = "% d-spacing"
                elif base_name == 'gamma':
                    cbar_label = "% FWHM"
                elif base_name == 'sigma':
                    cbar_label = "% Peak Width"
                elif base_name == 'area':
                    cbar_label = "% Peak Area"
                elif base_name == 'pos':
                    cbar_label = "% Peak Position"
                elif base_name == 'strain':
                    cbar_label = "% Strain"
                else:
                    cbar_label = f"% {base_name.capitalize()}"
            elif measurement == 'gamma':
                cbar_label = "FWHM (centidegrees)"
            elif measurement == 'strain':
                cbar_label = "Strain"
            elif measurement == 'd':
                cbar_label = "d-spacing (Å)"
            elif measurement == 'area':
                cbar_label = "Peak Area"
            elif measurement == 'sigma':
                cbar_label = "Peak Width"
            elif measurement == 'pos':
                cbar_label = "Peak Position"
            else:
                cbar_label = measurement.capitalize()

            # Create title - use analysis name if set, otherwise auto-generate
            tab_name = params['tab_name'].strip()
            if tab_name and tab_name != "Analysis" and not tab_name.startswith("Analysis "):
                # User has set a custom title
                title = tab_name
            else:
                # Auto-generate title: Peak Name Stage (Miller Index)
                peak_name = "Unknown Peak"
                stage_name = "Continuous"
                miller_index = "(hkl)"

                if hasattr(self.current_dataset.params, 'active_peaks') and peak_idx < len(self.current_dataset.params.active_peaks):
                    peak = self.current_dataset.params.active_peaks[peak_idx]
                    peak_name = peak.name
                    miller_index = f"({peak.miller_index})"

                if hasattr(self.current_dataset.params, 'stage'):
                    stage_map = {'BEF': 'Before', 'AFT': 'After', 'CONT': 'Continuous', 'DELT': 'Delta'}
                    stage_name = stage_map.get(self.current_dataset.params.stage.name, self.current_dataset.params.stage.name)

                title = f"{peak_name} {stage_name} {miller_index}"

            # Get azimuth range and spacing from dataset parameters
            azimuth_range = self.current_dataset.params.azimuths if hasattr(self.current_dataset.params, 'azimuths') else (-110, 110)
            spacing = self.current_dataset.params.spacing if hasattr(self.current_dataset.params, 'spacing') else 5

            # Use actual azimuth range from data if available
            if hasattr(df, 'columns') and len(df.columns) > 0:
                actual_az_min = float(df.columns.min())
                actual_az_max = float(df.columns.max())
                # Use actual range for axis formatting
                azimuth_range = (actual_az_min, actual_az_max)

            # Store current plot data for selection box functionality
            self.current_plot_data = df

            # Plot with corrected axis labels and DataFrame
            self.plot_canvas.plot_heatmap(
                df,
                xlabel="Azimuthal Angle (degrees)",  # Now x-axis (bottom)
                ylabel=ylabel,                       # Time/Depth on y-axis (side)
                colormap=params['colormap'],
                vmin=vmin,
                vmax=vmax,
                cbar_label=cbar_label,
                measurement_type=measurement,
                azimuth_range=azimuth_range,
                spacing=spacing,
                title=title,
                is_time_based=is_in_situ  # Pass the correct time/depth flag
            )

            # Add to gallery if generating
            if mode == "generate":
                self.add_to_gallery(params, title)

        except Exception as e:
            QMessageBox.critical(self, "Plot Error", f"Failed to create plot:\\n{str(e)}")

    def add_to_gallery(self, params, title):
        """Add generated plot to gallery."""
        # Create timestamp for unique identification
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Store plot info for gallery
        plot_info = {
            'title': title,
            'params': params.copy(),
            'timestamp': timestamp,
            'thumbnail': None  # TODO: Generate thumbnail
        }

        self.images_gallery.append(plot_info)

        # Update gallery display
        self.update_gallery_display()

    def export_csv(self):
        """Export data from the currently active analysis tab to CSV."""
        if not self.current_dataset:
            QMessageBox.warning(self, "No Dataset", "Please select a dataset first")
            return

        # Get current analysis tab
        current_tab = self.analysis_tabs.currentWidget()
        if not current_tab:
            QMessageBox.warning(self, "No Analysis Tab", "No analysis tab is currently active")
            return

        try:
            # Get parameters from current tab
            params = current_tab.get_analysis_parameters()
            measurement = params['measurement']
            peak_idx = params['peak_idx']

            # Get output file
            default_name = f"{self.current_dataset.params.sample}_{measurement}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export CSV", default_name, "CSV Files (*.csv)"
            )

            if not file_path:
                return

            # Handle delta/abs/pct measurements similar to plot_heatmap method
            is_delta = measurement.startswith('delta ')
            is_abs = measurement.startswith('abs ')
            is_pct = measurement.startswith('pct ')

            if measurement in self.current_dataset.col_idx:
                measurement_idx = self.current_dataset.col_idx[measurement]
            elif is_delta:
                base_measurement = measurement.replace('delta ', '')
                if base_measurement in self.current_dataset.col_idx:
                    self.current_dataset.calculate_delta(base_measurement)
                    measurement_idx = self.current_dataset.col_idx[measurement]
                else:
                    QMessageBox.warning(self, "Invalid Measurement",
                                      f"Base measurement '{base_measurement}' not found for delta calculation")
                    return
            elif is_abs:
                base_measurement = measurement.replace('abs ', '')
                if f"abs {base_measurement}" in self.current_dataset.col_idx:
                    measurement_idx = self.current_dataset.col_idx[f"abs {base_measurement}"]
                elif base_measurement in self.current_dataset.col_idx:
                    base_idx = self.current_dataset.col_idx[base_measurement]
                    base_data = self.current_dataset.data[:, :, :, base_idx]
                    if hasattr(base_data, 'compute'):
                        base_data = base_data.compute()
                    abs_data = np.abs(base_data)
                    self.current_dataset.add_measurement(f"abs {base_measurement}", abs_data)
                    measurement_idx = self.current_dataset.col_idx[f"abs {base_measurement}"]
                else:
                    QMessageBox.warning(self, "Invalid Measurement",
                                      f"Base measurement '{base_measurement}' not found for abs calculation")
                    return
            elif is_pct:
                base_measurement = measurement.replace('pct ', '')
                if base_measurement in self.current_dataset.col_idx:
                    # Calculate pct if it doesn't exist
                    self.current_dataset.calculate_pct(base_measurement)
                    measurement_idx = self.current_dataset.col_idx[measurement]
                else:
                    QMessageBox.warning(self, "Invalid Measurement",
                                      f"Base measurement '{base_measurement}' not found for pct calculation")
                    return
            else:
                QMessageBox.warning(self, "Invalid Measurement",
                                  f"Measurement '{measurement}' not found in dataset")
                return

            # Use same data preparation as plotting for consistency
            data_slice = self.current_dataset.data[peak_idx, :, :, measurement_idx]
            if hasattr(data_slice, 'compute'):
                data_slice = data_slice.compute()

            # Get frame and azimuth arrays for proper indexing
            frames = self.current_dataset.frame_numbers[peak_idx, :].compute()
            azimuths = self.current_dataset.azimuth_angles[peak_idx, :].compute()

            # Remove any zero azimuths (unfilled slots)
            valid_az_mask = azimuths != 0
            azimuths = azimuths[valid_az_mask]
            data_slice = data_slice[:, valid_az_mask]

            # Remove any zero frames (unfilled slots)
            valid_frame_mask = frames != 0
            if len(frames) > 0 and frames[0] == 0:
                valid_frame_mask[0] = True
            frames = frames[valid_frame_mask]
            data_slice = data_slice[valid_frame_mask, :]

            # Apply range filtering if specified
            if params.get('range_type') != 'full' and params.get('start_frame') is not None and params.get('end_frame') is not None:
                start_frame = params['start_frame']
                end_frame = params['end_frame']

                # Create range mask
                range_mask = (frames >= start_frame) & (frames <= end_frame)

                # Apply range mask
                frames = frames[range_mask]
                data_slice = data_slice[range_mask, :]

            # Convert frame numbers based on stage type
            is_in_situ = self.current_dataset.params.stage.name == 'CONT'
            if is_in_situ:
                time_values = (frames * 0.02) / 60
                df_index = time_values
            else:
                depth_values = np.arange(len(frames)-1, -1, -1)
                df_index = depth_values

            # Create DataFrame with proper orientation and labels
            df = pd.DataFrame(data_slice, index=df_index, columns=azimuths)
            df = df.loc[:, (df != 0).any(axis=0)]
            df = df.loc[(df != 0).any(axis=1)]
            df = df.sort_index()
            df = df.reindex(sorted(df.columns), axis=1)

            # Save as CSV with proper labels
            df.to_csv(file_path)
            QMessageBox.information(self, "Success", f"Data exported to:\\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export CSV:\\n{str(e)}")

    def save_plot(self):
        """Save the current plot to file."""
        if not self.plot_canvas.fig.axes:
            QMessageBox.warning(self, "No Plot", "Create a plot first")
            return

        try:
            # Get output file
            default_name = f"plot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Plot", default_name, "PNG Files (*.png);;PDF Files (*.pdf)"
            )

            if not file_path:
                return

            # Save plot
            self.plot_canvas.fig.savefig(file_path, dpi=300, bbox_inches='tight')
            QMessageBox.information(self, "Success", f"Plot saved to:\\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save plot:\\n{str(e)}")

    def clear_gallery(self):
        """Clear all items from the gallery."""
        self.images_gallery.clear()
        self.update_gallery_display()

    def set_grid_view(self):
        """Set gallery to grid view."""
        self.gallery_view_mode = "grid"
        self.grid_view_btn.setStyleSheet("QPushButton { background-color: #e8f4f8; font-weight: bold; }")
        self.carousel_view_btn.setStyleSheet("")
        self.update_gallery_display()

    def set_carousel_view(self):
        """Set gallery to carousel view."""
        self.gallery_view_mode = "carousel"
        self.carousel_view_btn.setStyleSheet("QPushButton { background-color: #e8f4f8; font-weight: bold; }")
        self.grid_view_btn.setStyleSheet("")
        self.update_gallery_display()

    def update_gallery_display(self):
        """Update the gallery display based on current mode."""
        # Clear existing widgets
        for i in reversed(range(self.gallery_layout.count())):
            child = self.gallery_layout.itemAt(i).widget()
            if child:
                child.setParent(None)

        if not self.images_gallery:
            # Show empty state
            empty_label = QLabel("No plots generated yet. Create plots using the analysis tabs.")
            empty_label.setStyleSheet("color: #666; font-style: italic; padding: 20px;")
            empty_label.setAlignment(Qt.AlignCenter)
            self.gallery_layout.addWidget(empty_label, 0, 0)
            return

        if self.gallery_view_mode == "grid":
            # Grid layout: show thumbnails in rows/columns
            cols = 3
            for i, plot_info in enumerate(self.images_gallery):
                row = i // cols
                col = i % cols

                # Create thumbnail widget
                thumbnail = self.create_thumbnail_widget(plot_info, i)
                self.gallery_layout.addWidget(thumbnail, row, col)

        else:  # carousel mode
            # Carousel layout: horizontal scrolling
            for i, plot_info in enumerate(self.images_gallery):
                thumbnail = self.create_thumbnail_widget(plot_info, i)
                self.gallery_layout.addWidget(thumbnail, 0, i)

    def create_thumbnail_widget(self, plot_info, index):
        """Create a thumbnail widget for a plot."""
        container = QWidget()
        container.setMaximumWidth(150)
        container.setMaximumHeight(120)
        layout = QVBoxLayout(container)

        # Title label (truncated)
        title = plot_info['title']
        if len(title) > 30:
            title = title[:27] + "..."

        title_label = QLabel(title)
        title_label.setWordWrap(True)
        title_label.setStyleSheet("font-size: 10px; font-weight: bold;")

        # Thumbnail placeholder (since we don't have actual image data)
        thumbnail_label = QLabel("📊")
        thumbnail_label.setAlignment(Qt.AlignCenter)
        thumbnail_label.setStyleSheet("""
            QLabel {
                border: 1px solid #ccc;
                background-color: #f5f5f5;
                font-size: 24px;
                min-height: 60px;
            }
        """)

        # Info label
        timestamp = plot_info['timestamp']
        measurement = plot_info['params']['measurement']
        info_label = QLabel(f"{measurement}\n{timestamp}")
        info_label.setStyleSheet("font-size: 8px; color: #666;")
        info_label.setAlignment(Qt.AlignCenter)

        # Make clickable
        thumbnail_label.mousePressEvent = lambda event, idx=index: self.on_thumbnail_clicked(idx)
        thumbnail_label.setCursor(Qt.PointingHandCursor)

        layout.addWidget(title_label)
        layout.addWidget(thumbnail_label)
        layout.addWidget(info_label)
        layout.setContentsMargins(5, 5, 5, 5)

        return container

    def on_thumbnail_clicked(self, index):
        """Handle thumbnail click - regenerate the plot."""
        if 0 <= index < len(self.images_gallery):
            plot_info = self.images_gallery[index]
            # Regenerate the plot with stored parameters
            self.plot_heatmap(plot_info['params'], "preview")

    def draw_selection_on_current_plot(self, frame_start, frame_end, azimuth_start, azimuth_end):
        """Draw selection box on the current plot."""
        if self.current_plot_data is not None:
            self.plot_canvas.draw_selection_box(frame_start, frame_end, azimuth_start, azimuth_end, self.current_plot_data)

    def update_all_tabs_with_dataset(self, dataset):
        """Update all analysis tabs with a new dataset."""
        # Set as current dataset
        self.current_dataset = dataset

        # Update all existing analysis tabs
        for i in range(self.analysis_tabs.count()):
            tab = self.analysis_tabs.widget(i)
            if tab:
                tab.update_for_dataset(dataset)

        # Update main UI
        self.update_ui_state()

    def run_batch_analysis(self):
        """Run automated batch analysis for all measurements and azimuth ranges."""
        if not self.current_dataset:
            QMessageBox.warning(self, "No Dataset", "Please load a dataset first.")
            return

        try:
            # Get parameters
            frame_start = self.batch_start_frame_spin.value()
            frame_end = self.batch_end_frame_spin.value()

            # Get azimuth ranges from UI
            if self.azimuth_range_data:
                range_info = self.azimuth_range_data
            else:
                # Default ranges if none defined
                range_info = [
                    {'name': 'Range_1', 'lower': -5, 'upper': 5},
                    {'name': 'Range_2', 'lower': 85, 'upper': 95},
                    {'name': 'Range_3', 'lower': 175, 'upper': 185},
                    {'name': 'Range_4', 'lower': -95, 'upper': -85}
                ]

            # Get all available measurements
            measurements = list(self.current_dataset.col_idx.keys())

            # Get all peaks
            n_peaks = self.current_dataset.data.shape[0]

            # Calculate total operations for progress bar
            total_operations = n_peaks * len(measurements) * len(range_info)

            # Setup progress bar
            self.batch_progress.setVisible(True)
            self.batch_progress.setMaximum(total_operations)
            self.batch_progress.setValue(0)

            # Clear previous results
            self.batch_results = []

            # Process each peak
            operation_count = 0
            for peak_idx in range(n_peaks):
                # Get peak name
                if (hasattr(self.current_dataset.params, 'active_peaks') and
                    self.current_dataset.params.active_peaks and
                    peak_idx < len(self.current_dataset.params.active_peaks)):
                    peak_name = self.current_dataset.params.active_peaks[peak_idx].name
                    peak_miller = self.current_dataset.params.active_peaks[peak_idx].miller_index
                else:
                    peak_name = f"Unknown {peak_idx+1}"
                    peak_miller = f"hkl{peak_idx+1}"

                # Process each measurement
                for measurement in measurements:
                    try:
                        # Get measurement data
                        measurement_idx = self.current_dataset.col_idx[measurement]
                        data_slice = self.current_dataset.data[peak_idx, :, :, measurement_idx].compute()
                        frames = self.current_dataset.frame_numbers[peak_idx, :].compute()
                        azimuths = self.current_dataset.azimuth_angles[peak_idx, :].compute()

                        # Clean data
                        valid_az_mask = azimuths != 0
                        azimuths = azimuths[valid_az_mask]
                        data_slice = data_slice[:, valid_az_mask]

                        valid_frame_mask = frames != 0
                        if len(frames) > 0 and frames[0] == 0:
                            valid_frame_mask[0] = True
                        frames = frames[valid_frame_mask]
                        data_slice = data_slice[valid_frame_mask, :]

                        # Process each azimuth range
                        for range_data in range_info:
                            az_start = range_data['lower']
                            az_end = range_data['upper']
                            range_name = range_data['name']
                            try:
                                # Extract range averages
                                avg_value = self.extract_range_average(
                                    data_slice, frames, azimuths,
                                    frame_start, frame_end, az_start, az_end
                                )

                                # Store result
                                self.batch_results.append({
                                    'Peak': peak_name,
                                    'Miller_Index': peak_miller,
                                    'Measurement': measurement,
                                    'Range_Name': range_name,
                                    'Azimuth_Range': f"({az_start},{az_end})",
                                    'Frame_Start': frame_start,
                                    'Frame_End': frame_end,
                                    'Average_Value': avg_value
                                })

                            except Exception as e:
                                print(f"Error processing {peak_name}-{measurement}-{range_name}({az_start},{az_end}): {e}")
                                # Store error result
                                self.batch_results.append({
                                    'Peak': peak_name,
                                    'Miller_Index': peak_miller,
                                    'Measurement': measurement,
                                    'Range_Name': range_name,
                                    'Azimuth_Range': f"({az_start},{az_end})",
                                    'Frame_Start': frame_start,
                                    'Frame_End': frame_end,
                                    'Average_Value': f"Error: {str(e)}"
                                })

                            # Update progress
                            operation_count += 1
                            self.batch_progress.setValue(operation_count)
                            QApplication.processEvents()  # Keep UI responsive

                    except Exception as e:
                        print(f"Error processing {peak_name}-{measurement}: {e}")
                        continue

            # Update UI
            self.batch_progress.setVisible(False)
            self.batch_results_label.setText(f"Results: {len(self.batch_results)} entries generated")
            self.export_csv_btn.setEnabled(True)
            self.export_excel_btn.setEnabled(True)

            QMessageBox.information(self, "Analysis Complete",
                                  f"Batch analysis complete!\n{len(self.batch_results)} entries generated.\nUse Export buttons to save results.")

        except Exception as e:
            self.batch_progress.setVisible(False)
            QMessageBox.critical(self, "Analysis Error", f"Error during batch analysis:\n{str(e)}")

    def extract_range_average(self, data_slice, frames, azimuths, frame_start, frame_end, az_start, az_end):
        """Extract average value for specific frame and azimuth range."""
        # Create masks
        frame_mask = (frames >= frame_start) & (frames <= frame_end)
        azimuth_mask = (azimuths >= az_start) & (azimuths <= az_end)

        # Extract selected data
        selected_data = data_slice[frame_mask, :][:, azimuth_mask]

        if selected_data.size == 0:
            return "No data in range"

        # Calculate average (exclude zeros)
        non_zero_data = selected_data[selected_data != 0]
        if len(non_zero_data) == 0:
            return "All values are zero"

        return float(np.mean(non_zero_data))

    def export_batch_csv(self):
        """Export batch analysis results to CSV."""
        if not self.batch_results:
            QMessageBox.warning(self, "No Results", "No batch analysis results to export.")
            return

        try:
            # Get save location
            filename, _ = QFileDialog.getSaveFileName(
                self, "Save Batch Analysis Results",
                f"batch_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "CSV Files (*.csv)"
            )

            if not filename:
                return

            # Create DataFrame and save
            import pandas as pd
            df = pd.DataFrame(self.batch_results)
            df.to_csv(filename, index=False)

            QMessageBox.information(self, "Export Complete", f"Results exported to:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export CSV:\n{str(e)}")

    def export_batch_excel(self):
        """Export batch analysis results to Excel using template format."""
        if not self.batch_results:
            QMessageBox.warning(self, "No Results", "No batch analysis results to export.")
            return

        try:
            # Get save location
            filename, _ = QFileDialog.getSaveFileName(
                self, "Save Batch Analysis Results",
                f"batch_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not filename:
                return

            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, Alignment
            import os

            # Organize data by peak and measurement type
            organized_data = self._organize_batch_data_for_template()

            # Create new workbook with simple template structure
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create single data sheet
            sheet = wb.create_sheet('Batch_Analysis_Data')
            self._create_simple_template_sheet(sheet, organized_data)

            # Save the workbook
            wb.save(filename)
            wb.close()

            QMessageBox.information(self, "Export Complete",
                                  f"Results exported to:\n{filename}\n\nSheet created: Batch_Analysis_Data\n\nData is formatted according to your template structure.")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n{str(e)}")
            # Fallback to CSV if Excel export fails
            reply = QMessageBox.question(self, "Excel Failed", "Excel export failed. Export as CSV instead?")
            if reply == QMessageBox.Yes:
                self.export_batch_csv()

    def _organize_batch_data_for_template(self):
        """Organize batch results into template-compatible structure."""
        import pandas as pd

        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame(self.batch_results)

        # Organize by peak
        organized = {}

        for peak in df['Peak'].unique():
            peak_data = df[df['Peak'] == peak]
            organized[peak] = {}

            # Get Miller index for this peak
            miller_index = peak_data['Miller_Index'].iloc[0] if len(peak_data) > 0 else 'hkl'
            organized[peak]['miller_index'] = miller_index

            # Organize by measurement type
            for measurement in peak_data['Measurement'].unique():
                measurement_data = peak_data[peak_data['Measurement'] == measurement]
                organized[peak][measurement] = {}

                # Organize by range name (azimuthal range)
                for _, row in measurement_data.iterrows():
                    range_name = row['Range_Name']
                    try:
                        avg_value = float(row['Average_Value'])
                    except (ValueError, TypeError):
                        avg_value = 0.0

                    organized[peak][measurement][range_name] = avg_value

        return organized

    def _parse_range_name(self, range_name):
        """Extract base name by removing parenthetical content."""
        import re
        # Remove content in parentheses: "(-) Normal" -> "Normal"
        base_name = re.sub(r'\([^)]*\)', '', range_name).strip()
        return base_name if base_name else range_name

    def _group_ranges_by_base(self, range_list):
        """Group ranges by their base names."""
        groups = {}
        for range_name in range_list:
            base_name = self._parse_range_name(range_name)
            if base_name not in groups:
                groups[base_name] = []
            groups[base_name].append(range_name)
        return groups

    def _calculate_column_layout(self, range_list):
        """Calculate the complete column layout with data, averages, and ratios."""
        groups = self._group_ranges_by_base(range_list)
        base_names = sorted(groups.keys())  # Sort for consistent ordering

        layout = {
            'data_columns': range_list,  # Original range names
            'average_columns': [f'Avg {base}' for base in base_names],
            'ratio_columns': []
        }

        # Create ratios
        if len(base_names) == 2:
            layout['ratio_columns'] = [f'{base_names[0]}/{base_names[1]}']
        elif len(base_names) > 2:
            for i in range(len(base_names)):
                for j in range(i+1, len(base_names)):
                    layout['ratio_columns'].append(f'{base_names[i]}/{base_names[j]}')

        return layout, groups

    def _create_average_formula(self, groups, base_name, current_row, range_to_col_mapping):
        """Create average formula for a base name group."""
        range_names = groups[base_name]
        col_refs = []

        for range_name in range_names:
            if range_name in range_to_col_mapping:
                col_refs.append(f'{range_to_col_mapping[range_name]}{current_row}')

        if col_refs:
            return f'=AVERAGE({",".join(col_refs)})'
        return 0

    def _create_ratio_formula(self, avg_col_mapping, ratio_name, current_row):
        """Create ratio formula between average columns."""
        parts = ratio_name.split('/')
        if len(parts) == 2:
            numerator_col = avg_col_mapping.get(f'Avg {parts[0]}')
            denominator_col = avg_col_mapping.get(f'Avg {parts[1]}')
            if numerator_col and denominator_col:
                return f'={numerator_col}{current_row}/{denominator_col}{current_row}'
        return 0

    def _create_simple_template_sheet(self, sheet, organized_data):
        """Create a simple template-formatted sheet with smart grouping and calculations."""
        from openpyxl.styles import Font, Alignment, PatternFill

        # Header styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        # Add title and sample info
        sheet['B1'] = 'Azimuthal Analysis Results'
        sheet['B1'].font = Font(bold=True, size=14)

        sheet['B2'] = 'Sample:'
        if hasattr(self.current_dataset, 'params') and hasattr(self.current_dataset.params, 'sample'):
            sheet['C2'] = self.current_dataset.params.sample

        # Get all range names and calculate smart layout
        all_ranges = set()
        for peak_data in organized_data.values():
            for measurement_data in peak_data.values():
                if isinstance(measurement_data, dict):
                    all_ranges.update(measurement_data.keys())

        range_list = sorted(all_ranges)
        layout, groups = self._calculate_column_layout(range_list)

        # Calculate total columns needed
        total_cols_needed = len(layout['data_columns']) + len(layout['average_columns']) + len(layout['ratio_columns'])
        base_cols = [chr(68 + i) for i in range(min(total_cols_needed, 20))]  # D through X (max 20 cols)

        # Create column mappings
        col_idx = 0
        range_to_col_mapping = {}
        avg_col_mapping = {}
        ratio_col_mapping = {}

        # Map data columns
        for range_name in layout['data_columns']:
            if col_idx < len(base_cols):
                range_to_col_mapping[range_name] = base_cols[col_idx]
                col_idx += 1

        # Map average columns
        for avg_name in layout['average_columns']:
            if col_idx < len(base_cols):
                avg_col_mapping[avg_name] = base_cols[col_idx]
                col_idx += 1

        # Map ratio columns
        for ratio_name in layout['ratio_columns']:
            if col_idx < len(base_cols):
                ratio_col_mapping[ratio_name] = base_cols[col_idx]
                col_idx += 1

        # Setup headers for each peak
        available_peaks = list(organized_data.keys())
        current_row = 4

        for peak_idx, peak_name in enumerate(available_peaks):
            peak_data = organized_data[peak_name]

            # Peak header
            sheet[f'B{current_row}'] = 'Peak:'
            sheet[f'B{current_row}'].font = header_font
            sheet[f'B{current_row}'].fill = header_fill

            miller_index = peak_data.get('miller_index', peak_name)
            sheet[f'C{current_row}'] = miller_index
            sheet[f'C{current_row}'].font = header_font

            # Data column headers
            for range_name in layout['data_columns']:
                if range_name in range_to_col_mapping:
                    col = range_to_col_mapping[range_name]
                    sheet[f'{col}{current_row}'] = range_name
                    sheet[f'{col}{current_row}'].font = header_font
                    sheet[f'{col}{current_row}'].fill = header_fill

            # Average column headers
            for avg_name in layout['average_columns']:
                if avg_name in avg_col_mapping:
                    col = avg_col_mapping[avg_name]
                    sheet[f'{col}{current_row}'] = avg_name
                    sheet[f'{col}{current_row}'].font = header_font
                    sheet[f'{col}{current_row}'].fill = header_fill

            # Ratio column headers
            for ratio_name in layout['ratio_columns']:
                if ratio_name in ratio_col_mapping:
                    col = ratio_col_mapping[ratio_name]
                    sheet[f'{col}{current_row}'] = ratio_name
                    sheet[f'{col}{current_row}'].font = header_font
                    sheet[f'{col}{current_row}'].fill = header_fill

            current_row += 1

            # Get actual measurement names from the peak data
            measurements = []
            if peak_name in organized_data:
                measurements = [key for key in organized_data[peak_name].keys() if key != 'miller_index']

            # Sort measurements for consistent ordering (put common ones first)
            measurement_priority = ['d', 'strain', 'area', 'pos', 'sigma', 'gamma']
            priority_measurements = [m for m in measurement_priority if m in measurements]
            other_measurements = [m for m in measurements if m not in measurement_priority]
            measurements = priority_measurements + sorted(other_measurements)


            for measurement in measurements:
                sheet[f'B{current_row}'] = measurement

                # Add data if available
                if measurement in peak_data:
                    measurement_data = peak_data[measurement]

                    # Fill data for each azimuthal range
                    for range_name in layout['data_columns']:
                        if range_name in range_to_col_mapping and range_name in measurement_data:
                            col = range_to_col_mapping[range_name]
                            try:
                                value = float(measurement_data[range_name])
                                sheet[f'{col}{current_row}'] = value
                            except (ValueError, TypeError):
                                sheet[f'{col}{current_row}'] = 0.0

                # Fill average formulas (outside the if block)
                for avg_name in layout['average_columns']:
                    if avg_name in avg_col_mapping:
                        col = avg_col_mapping[avg_name]
                        base_name = avg_name.replace('Avg ', '')
                        formula = self._create_average_formula(groups, base_name, current_row, range_to_col_mapping)
                        sheet[f'{col}{current_row}'] = formula

                # Fill ratio formulas
                for ratio_name in layout['ratio_columns']:
                    if ratio_name in ratio_col_mapping:
                        col = ratio_col_mapping[ratio_name]
                        formula = self._create_ratio_formula(avg_col_mapping, ratio_name, current_row)
                        sheet[f'{col}{current_row}'] = formula

                current_row += 1

            # Add spacing between peaks
            current_row += 1

        # Auto-adjust column widths
        for col_letter in base_cols[:col_idx]:
            sheet.column_dimensions[col_letter].width = 15

    def update_ranges_display(self):
        """Update the azimuth ranges list widget display."""
        self.azimuth_ranges_list.clear()
        for range_data in self.azimuth_range_data:
            item_text = f"{range_data['name']}: ({range_data['lower']}, {range_data['upper']})°"
            self.azimuth_ranges_list.addItem(item_text)

    def add_azimuth_range(self):
        """Add a new azimuth range."""
        dialog = AzimuthRangeDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            range_data = dialog.get_range_data()
            self.azimuth_range_data.append(range_data)
            self.update_ranges_display()
            self.schedule_settings_save()

    def edit_azimuth_range(self):
        """Edit the selected azimuth range."""
        current_row = self.azimuth_ranges_list.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "No Selection", "Please select a range to edit.")
            return

        current_data = self.azimuth_range_data[current_row]
        dialog = AzimuthRangeDialog(self, current_data)
        if dialog.exec_() == QDialog.Accepted:
            range_data = dialog.get_range_data()
            self.azimuth_range_data[current_row] = range_data
            self.update_ranges_display()
            self.schedule_settings_save()

    def delete_azimuth_range(self):
        """Delete the selected azimuth range."""
        current_row = self.azimuth_ranges_list.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "No Selection", "Please select a range to delete.")
            return

        range_data = self.azimuth_range_data[current_row]
        reply = QMessageBox.question(self, "Confirm Delete",
                                   f"Delete range '{range_data['name']}'?")
        if reply == QMessageBox.Yes:
            self.azimuth_range_data.pop(current_row)
            self.update_ranges_display()
            self.schedule_settings_save()

    def reset_default_ranges(self):
        """Reset to the default azimuth ranges."""
        reply = QMessageBox.question(self, "Reset Ranges",
                                   "Reset to default azimuth ranges? This will remove all custom ranges.")
        if reply == QMessageBox.Yes:
            self.azimuth_range_data = [
                {"name": "Near Loading", "lower": -5, "upper": 5},
                {"name": "Transverse 1", "lower": 85, "upper": 95},
                {"name": "Opposite Loading", "lower": 175, "upper": 185},
                {"name": "Transverse 2", "lower": -95, "upper": -85}
            ]
            self.update_ranges_display()
            self.schedule_settings_save()


class AzimuthRangeDialog(QDialog):
    """Dialog for adding or editing azimuth ranges."""

    def __init__(self, parent=None, range_data=None):
        super().__init__(parent)
        self.setWindowTitle("Define Azimuth Range")
        self.setModal(True)
        self.resize(300, 150)

        layout = QFormLayout(self)

        # Range name
        self.name_edit = QLineEdit()
        if range_data:
            self.name_edit.setText(range_data['name'])
        else:
            self.name_edit.setText("Custom Range")
        layout.addRow("Range Name:", self.name_edit)

        # Lower limit
        self.lower_spin = QDoubleSpinBox()
        self.lower_spin.setRange(-360.0, 360.0)
        self.lower_spin.setSingleStep(1.0)
        self.lower_spin.setSuffix("°")
        if range_data:
            self.lower_spin.setValue(range_data['lower'])
        else:
            self.lower_spin.setValue(-10.0)
        layout.addRow("Lower Limit:", self.lower_spin)

        # Upper limit
        self.upper_spin = QDoubleSpinBox()
        self.upper_spin.setRange(-360.0, 360.0)
        self.upper_spin.setSingleStep(1.0)
        self.upper_spin.setSuffix("°")
        if range_data:
            self.upper_spin.setValue(range_data['upper'])
        else:
            self.upper_spin.setValue(10.0)
        layout.addRow("Upper Limit:", self.upper_spin)

        # Buttons
        button_layout = QHBoxLayout()
        self.ok_btn = QPushButton("OK")
        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)

        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addRow("", button_layout)

        # Validation
        self.lower_spin.valueChanged.connect(self.validate_range)
        self.upper_spin.valueChanged.connect(self.validate_range)
        self.name_edit.textChanged.connect(self.validate_range)

        self.validate_range()

    def validate_range(self):
        """Validate the range inputs and enable/disable OK button."""
        name = self.name_edit.text().strip()
        lower = self.lower_spin.value()
        upper = self.upper_spin.value()

        valid = len(name) > 0 and upper > lower
        self.ok_btn.setEnabled(valid)

        if not valid and upper <= lower:
            self.upper_spin.setStyleSheet("QDoubleSpinBox { border: 2px solid red; }")
        else:
            self.upper_spin.setStyleSheet("")

    def get_range_data(self):
        """Get the range data from the dialog."""
        return {
            'name': self.name_edit.text().strip(),
            'lower': self.lower_spin.value(),
            'upper': self.upper_spin.value()
        }


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Modern look

    window = DataAnalyzer()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()